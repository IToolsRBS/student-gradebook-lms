"""
Build Activity Completion export workbook from warehouse gradebook marts.

Filters (CLI; empty / omitted = select all for that dimension):
  category, programme, module, assessment type, assessment,
  mark status (graded / not_graded).

Always restricted to submitted assessments (status submitted or submitted_late).
Mark-status filters apply only within those submitted rows — missed is excluded.

Sheets:
  1. Activity Completion Summary — one row per assessment
  2. Submission Details — row per matching submission (like Student Assessment Detail)

Prints the absolute output path as the last stdout line (for future frontend wiring).
"""

from __future__ import annotations

import argparse
import gc
from datetime import datetime
from pathlib import Path
from typing import Any, Iterator, Sequence

import duckdb
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from motherduck_client import (
    connect_motherduck,
    gradebook_schema,
    qualified_relation,
)
from populate_gradebook_from_warehouse import (
    COLUMN_WIDTH_SAMPLE_ROWS,
    DEFAULT_SCHEMA,
    FETCH_CHUNK_SIZE,
    MAX_COLUMN_WIDTH,
    NOTE_FIELD_MAP,
    TABLE_ASSESSMENT,
    TABLE_MODULE,
    _mart_columns,
    _pick_first_mart_column,
    _update_col_widths,
    append_data_row,
    finish_sheet,
    format_cell,
    normalize_row,
    pick,
    write_headers,
)

ALLOWED_STATUSES: tuple[str, ...] = ("submitted", "graded", "submitted_late")
# Activity completion always includes only these submission statuses (never missed).
SUBMISSION_STATUSES: tuple[str, ...] = ("submitted", "submitted_late")
MARK_STATUSES: tuple[str, ...] = ("graded", "not_graded")

SUMMARY_HEADERS: list[str] = [
    "Category",
    "Programme",
    "Module Code",
    "Module",
    "Assessment",
    "Assessment Type",
    "Total Students In Module",
    "Total Students Suspended In Module",
    "Total Submissions",
]

DETAIL_HEADERS: list[str] = [
    "Category",
    "Programme",
    "Student No",
    "Student",
    "Email",
    "Module Code",
    "Module",
    "Assessment",
    "Assessment Type",
    "Due Date",
    "Submitted Date",
    "Status",
    "Mark Status",
    "Mark",
    "Max Grade",
    *[label for label, _ in NOTE_FIELD_MAP],
]


def _normalize_filter_values(values: Sequence[str] | None) -> list[str]:
    """Dedupe trimmed values; empty input means select-all."""
    seen: set[str] = set()
    out: list[str] = []
    for raw in values or []:
        text = str(raw).strip()
        if not text or text == "*":
            continue
        key = text.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(text)
    return out


def _normalize_mark_statuses(values: Sequence[str] | None) -> list[str]:
    """
    Mark-status filter for activity completion (graded / not_graded).
    Empty / select-all → no mark_status filter (return empty list).
    """
    requested = _normalize_filter_values(values)
    if not requested:
        return []

    allowed_lookup = {s: s for s in MARK_STATUSES}
    # Accept "not graded" → not_graded
    resolved: list[str] = []
    seen: set[str] = set()
    invalid: list[str] = []
    for raw in requested:
        key = raw.strip().lower().replace(" ", "_").replace("-", "_")
        canonical = allowed_lookup.get(key)
        if not canonical:
            invalid.append(raw)
            continue
        if canonical not in seen:
            seen.add(canonical)
            resolved.append(canonical)
    if invalid:
        raise SystemExit(
            "Invalid --status value(s): "
            + ", ".join(invalid)
            + f". Allowed mark statuses: {', '.join(MARK_STATUSES)}"
        )
    return resolved


def _normalize_statuses(values: Sequence[str] | None) -> list[str]:
    """
    Return allowed completion statuses to include.
    Empty / select-all → all ALLOWED_STATUSES.
    Unknown values are rejected.
    """
    requested = _normalize_filter_values(values)
    if not requested:
        return list(ALLOWED_STATUSES)

    allowed_lookup = {s: s for s in ALLOWED_STATUSES}
    resolved: list[str] = []
    seen: set[str] = set()
    invalid: list[str] = []
    for raw in requested:
        key = raw.strip().lower().replace(" ", "_").replace("-", "_")
        canonical = allowed_lookup.get(key)
        if not canonical:
            invalid.append(raw)
            continue
        if canonical not in seen:
            seen.add(canonical)
            resolved.append(canonical)
    if invalid:
        raise SystemExit(
            "Invalid --status value(s): "
            + ", ".join(invalid)
            + f". Allowed: {', '.join(ALLOWED_STATUSES)}"
        )
    return resolved or list(ALLOWED_STATUSES)


def _in_clause(column_sql: str, values: Sequence[str], params: list[Any]) -> str:
    placeholders = ", ".join("?" for _ in values)
    params.extend(values)
    return f"{column_sql} IN ({placeholders})"


def _truthy_sql(column_name: str) -> str:
    """SQL predicate: column is a boolean/string/number truthy flag."""
    col = f'"{column_name}"'
    return (
        f"(COALESCE(TRY_CAST({col} AS BOOLEAN), false) = true "
        f"OR LOWER(TRIM(CAST({col} AS VARCHAR))) IN ('true', 't', '1', 'yes', 'y'))"
    )


def _submitted_branch_sql(mart_cols: dict[str, str]) -> str:
    """Return SQL expr that yields 'submitted' or 'submitted_late' from due/attempt times."""
    due_col = _pick_first_mart_column(
        mart_cols, ("due_at", "effective_deadline_at", "due_date")
    )
    attempt_at_col = _pick_first_mart_column(
        mart_cols, ("last_attempt_at", "grade_submitted_at", "submitted_at", "graded_at")
    )
    if due_col and attempt_at_col:
        late_pred = (
            f'("{due_col}" IS NOT NULL AND "{attempt_at_col}" IS NOT NULL '
            f'AND CAST("{attempt_at_col}" AS TIMESTAMP) > CAST("{due_col}" AS TIMESTAMP))'
        )
        return f"CASE WHEN {late_pred} THEN 'submitted_late' ELSE 'submitted' END"
    return "'submitted'"


def _was_submitted_sql(mart_cols: dict[str, str]) -> str:
    """SQL predicate: row has submit/attempt evidence."""
    is_submitted_col = _pick_first_mart_column(mart_cols, ("is_submitted",))
    has_attempt_col = _pick_first_mart_column(mart_cols, ("has_attempt",))
    attempt_count_col = _pick_first_mart_column(mart_cols, ("attempt_count",))
    submitted_parts: list[str] = []
    if is_submitted_col:
        submitted_parts.append(_truthy_sql(is_submitted_col))
    if has_attempt_col:
        submitted_parts.append(_truthy_sql(has_attempt_col))
    if attempt_count_col:
        submitted_parts.append(
            f'(TRY_CAST("{attempt_count_col}" AS BIGINT) IS NOT NULL '
            f'AND TRY_CAST("{attempt_count_col}" AS BIGINT) > 0)'
        )
    return " OR ".join(submitted_parts) if submitted_parts else "false"


def _effective_status_sql(mart_cols: dict[str, str]) -> str | None:
    """
    Derive submission status for filtering/display: submitted or submitted_late.

    Warehouse may store status='graded' for a submitted+graded row — map that to
    submitted(_late). Do not treat missed-with-a-grade as submitted; missed stays
    missed and is excluded from activity completion.
    """
    status_col = _pick_first_mart_column(mart_cols, ("status",))
    if not status_col:
        return None

    raw_status = (
        "LOWER(REPLACE(REPLACE(TRIM(CAST("
        f'"{status_col}" AS VARCHAR)), \' \', \'_\'), \'-\', \'_\'))'
    )
    was_submitted = _was_submitted_sql(mart_cols)
    submitted_branch = _submitted_branch_sql(mart_cols)

    return f"""
        CASE
            WHEN {raw_status} IN ('submitted', 'submitted_late')
                THEN {raw_status}
            WHEN {raw_status} = 'graded' OR ({was_submitted})
                THEN {submitted_branch}
            ELSE {raw_status}
        END
    """.strip()


def _submission_lateness(row: dict[str, Any]) -> str:
    """Return 'submitted_late' or 'submitted' from due/attempt timestamps."""
    due = pick(row, "due_at", "effective_deadline_at", "due_date")
    attempt_at = pick(
        row, "last_attempt_at", "grade_submitted_at", "submitted_at", "graded_at"
    )
    if due and attempt_at:
        try:
            due_dt = due if isinstance(due, datetime) else datetime.fromisoformat(str(due))
            att_dt = (
                attempt_at
                if isinstance(attempt_at, datetime)
                else datetime.fromisoformat(str(attempt_at))
            )
            if getattr(due_dt, "tzinfo", None):
                due_dt = due_dt.replace(tzinfo=None)
            if getattr(att_dt, "tzinfo", None):
                att_dt = att_dt.replace(tzinfo=None)
            if att_dt > due_dt:
                return "submitted_late"
        except (TypeError, ValueError, OSError):
            pass
    return "submitted"


def effective_completion_status(row: dict[str, Any]) -> str:
    """Python-side mirror of _effective_status_sql for display."""
    raw = (
        str(pick(row, "effective_status", "status") or "")
        .strip()
        .lower()
        .replace(" ", "_")
        .replace("-", "_")
    )
    if raw in SUBMISSION_STATUSES:
        return raw

    def _flag(value: Any) -> bool:
        if value in (True, 1):
            return True
        return str(value).strip().lower() in {"true", "t", "1", "yes", "y"}

    was_submitted = (
        raw == "graded"
        or _flag(row.get("is_submitted"))
        or _flag(row.get("has_attempt"))
        or (
            str(pick(row, "attempt_count") or "").strip().isdigit()
            and int(pick(row, "attempt_count")) > 0
        )
    )
    if was_submitted:
        return _submission_lateness(row)
    return raw


def _build_dimension_where(
    mart_cols: dict[str, str],
    *,
    category_names: Sequence[str],
    programme_codes: Sequence[str],
    modules: Sequence[str],
    assessment_types: Sequence[str],
    assessments: Sequence[str],
) -> tuple[list[str], list[Any]]:
    programme_col = _pick_first_mart_column(
        mart_cols, ("programme", "course_prefix", "program_code")
    )
    category_col = _pick_first_mart_column(mart_cols, ("category_name",))
    module_col = _pick_first_mart_column(
        mart_cols, ("course_shortname", "module_code", "course_code")
    )
    assessment_type_col = _pick_first_mart_column(mart_cols, ("assessment_type",))
    assessment_col = _pick_first_mart_column(
        mart_cols, ("assessment", "assessment_name")
    )

    where_parts: list[str] = []
    params: list[Any] = []

    if category_names and category_col:
        where_parts.append(
            _in_clause(
                f'TRIM(CAST("{category_col}" AS VARCHAR))',
                [c.strip() for c in category_names],
                params,
            )
        )
    if programme_codes and programme_col:
        where_parts.append(
            _in_clause(
                f'UPPER(TRIM(CAST("{programme_col}" AS VARCHAR)))',
                [c.strip().upper() for c in programme_codes],
                params,
            )
        )
    if modules and module_col:
        where_parts.append(
            _in_clause(
                f'UPPER(TRIM(CAST("{module_col}" AS VARCHAR)))',
                [m.strip().upper() for m in modules],
                params,
            )
        )
    if assessment_types and assessment_type_col:
        where_parts.append(
            _in_clause(
                f'UPPER(TRIM(CAST("{assessment_type_col}" AS VARCHAR)))',
                [t.strip().upper() for t in assessment_types],
                params,
            )
        )
    if assessments and assessment_col:
        where_parts.append(
            _in_clause(
                f'UPPER(TRIM(CAST("{assessment_col}" AS VARCHAR)))',
                [a.strip().upper() for a in assessments],
                params,
            )
        )

    return where_parts, params


def _build_activity_filter_sql(
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    *,
    category_names: Sequence[str],
    programme_codes: Sequence[str],
    modules: Sequence[str],
    assessment_types: Sequence[str],
    assessments: Sequence[str],
    statuses: Sequence[str],
    mark_statuses: Sequence[str] | None = None,
    due_from: str | None = None,
    due_to: str | None = None,
    order_columns: Sequence[str] | None = None,
    select_sql: str = "*",
    group_by_sql: str | None = None,
) -> tuple[str, list[Any]] | None:
    """Build filtered SELECT against gradebook_student_assessment_detail.

    When ``mark_statuses`` is provided (activity completion), rows are always
    restricted to submitted / submitted_late, and ``mark_statuses`` further
    filters mark_status. ``statuses`` is ignored in that mode.

    When ``mark_statuses`` is None (late submissions), ``statuses`` filters
    effective completion status as before.
    """
    relation = qualified_relation(schema, TABLE_ASSESSMENT)
    try:
        mart_cols = _mart_columns(conn, schema, TABLE_ASSESSMENT)
    except duckdb.CatalogException:
        return None

    dim_where, params = _build_dimension_where(
        mart_cols,
        category_names=category_names,
        programme_codes=programme_codes,
        modules=modules,
        assessment_types=assessment_types,
        assessments=assessments,
    )

    due_col = _pick_first_mart_column(
        mart_cols, ("due_at", "effective_deadline_at", "due_date")
    )
    if due_from and due_col:
        dim_where.append(
            f'TRY_CAST("{due_col}" AS DATE) >= TRY_CAST(? AS DATE)'
        )
        params.append(due_from)
    if due_to and due_col:
        dim_where.append(
            f'TRY_CAST("{due_col}" AS DATE) <= TRY_CAST(? AS DATE)'
        )
        params.append(due_to)

    base_where = " AND ".join(dim_where) if dim_where else "1=1"

    effective_sql = _effective_status_sql(mart_cols)

    if mark_statuses is not None:
        # Activity completion: always submitted / submitted_late only.
        if not effective_sql:
            return None
        mark_status_values = [
            s.strip().lower().replace(" ", "_").replace("-", "_")
            for s in mark_statuses
            if str(s).strip()
        ]
        filter_params: list[Any] = []
        submission_clause = _in_clause(
            "effective_status",
            list(SUBMISSION_STATUSES),
            filter_params,
        )
        where_extra = submission_clause
        if mark_status_values:
            mark_col = _pick_first_mart_column(mart_cols, ("mark_status",))
            if not mark_col:
                return None
            mark_clause = _in_clause(
                f'LOWER(REPLACE(REPLACE(TRIM(CAST("{mark_col}" AS VARCHAR)), '
                f"' ', '_'), '-', '_'))",
                mark_status_values,
                filter_params,
            )
            where_extra = f"{submission_clause} AND {mark_clause}"
        inner = (
            f'SELECT *, ({effective_sql}) AS effective_status '
            f"FROM {relation} WHERE {base_where}"
        )
        query = f"SELECT {select_sql} FROM ({inner}) AS _activity WHERE {where_extra}"
        params = list(params) + filter_params
    elif statuses:
        if not effective_sql:
            return None
        status_params: list[Any] = []
        status_clause = _in_clause(
            "effective_status",
            [str(s).strip().lower().replace(" ", "_").replace("-", "_") for s in statuses],
            status_params,
        )
        inner = (
            f'SELECT *, ({effective_sql}) AS effective_status '
            f"FROM {relation} WHERE {base_where}"
        )
        query = f"SELECT {select_sql} FROM ({inner}) AS _activity WHERE {status_clause}"
        params = list(params) + status_params
    else:
        query = f"SELECT {select_sql} FROM {relation} WHERE {base_where}"

    if group_by_sql:
        query += f" GROUP BY {group_by_sql}"
    if order_columns:
        present = [col for col in order_columns if col.lower() in mart_cols]
        if present:
            order = ", ".join(f'"{mart_cols[col.lower()]}"' for col in present)
            query += f" ORDER BY {order}"
        elif not group_by_sql:
            query += " ORDER BY 1"
    elif not group_by_sql:
        query += " ORDER BY 1"
    return query, params


def iter_filtered_assessment_rows(
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    *,
    category_names: Sequence[str],
    programme_codes: Sequence[str],
    modules: Sequence[str],
    assessment_types: Sequence[str],
    assessments: Sequence[str],
    statuses: Sequence[str],
    mark_statuses: Sequence[str] | None = None,
    due_from: str | None = None,
    due_to: str | None = None,
    order_columns: Sequence[str] | None = None,
    chunk_size: int = FETCH_CHUNK_SIZE,
) -> Iterator[dict[str, Any]]:
    built = _build_activity_filter_sql(
        conn,
        schema,
        category_names=category_names,
        programme_codes=programme_codes,
        modules=modules,
        assessment_types=assessment_types,
        assessments=assessments,
        statuses=statuses,
        mark_statuses=mark_statuses,
        due_from=due_from,
        due_to=due_to,
        order_columns=order_columns,
    )
    if not built:
        return
    query, params = built
    result = conn.execute(query, params)
    columns = [str(desc[0]) for desc in result.description]
    while True:
        batch = result.fetchmany(chunk_size)
        if not batch:
            break
        for tup in batch:
            yield normalize_row(dict(zip(columns, tup)))


def _module_stats_lookup(
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    *,
    category_names: Sequence[str],
    programme_codes: Sequence[str],
    modules: Sequence[str],
) -> dict[tuple[str, str], dict[str, Any]]:
    """
    Map (programme, module_code) → {students, suspended, module_name, category}.
    Uses gradebook_module_summary when available.
    """
    relation = qualified_relation(schema, TABLE_MODULE)
    try:
        mart_cols = _mart_columns(conn, schema, TABLE_MODULE)
    except duckdb.CatalogException:
        return {}

    programme_col = _pick_first_mart_column(
        mart_cols, ("programme", "program_code")
    )
    module_code_col = _pick_first_mart_column(
        mart_cols, ("module_code", "course_shortname")
    )
    module_name_col = _pick_first_mart_column(mart_cols, ("module", "course_fullname"))
    students_col = _pick_first_mart_column(mart_cols, ("students", "total_students"))
    suspended_col = _pick_first_mart_column(
        mart_cols,
        ("suspended_students", "suspended_student_count", "suspended", "students_suspended"),
    )
    category_col = _pick_first_mart_column(mart_cols, ("category_name",))
    if not programme_col or not module_code_col:
        return {}

    where_parts: list[str] = ["1=1"]
    params: list[Any] = []
    if category_names and category_col:
        where_parts.append(
            _in_clause(
                f'TRIM(CAST("{category_col}" AS VARCHAR))',
                [c.strip() for c in category_names],
                params,
            )
        )
    if programme_codes:
        where_parts.append(
            _in_clause(
                f'UPPER(TRIM(CAST("{programme_col}" AS VARCHAR)))',
                [c.strip().upper() for c in programme_codes],
                params,
            )
        )
    if modules:
        where_parts.append(
            _in_clause(
                f'UPPER(TRIM(CAST("{module_code_col}" AS VARCHAR)))',
                [m.strip().upper() for m in modules],
                params,
            )
        )

    select_parts = [
        f'"{programme_col}" AS programme',
        f'"{module_code_col}" AS module_code',
    ]
    if module_name_col:
        select_parts.append(f'"{module_name_col}" AS module_name')
    if students_col:
        select_parts.append(f'"{students_col}" AS students')
    if suspended_col:
        select_parts.append(f'"{suspended_col}" AS suspended_students')
    if category_col:
        select_parts.append(f'"{category_col}" AS category_name')

    query = (
        f"SELECT {', '.join(select_parts)} FROM {relation} "
        f"WHERE {' AND '.join(where_parts)}"
    )
    try:
        result = conn.execute(query, params)
        columns = [str(d[0]).lower() for d in result.description]
        rows = result.fetchall()
    except Exception:
        return {}

    lookup: dict[tuple[str, str], dict[str, Any]] = {}
    for tup in rows:
        row = normalize_row(dict(zip(columns, tup)))
        programme = str(pick(row, "programme") or "").strip().upper()
        module_code = str(pick(row, "module_code") or "").strip().upper()
        if not programme or not module_code:
            continue
        lookup[(programme, module_code)] = row
    return lookup


def _suspended_by_module_from_detail(
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    *,
    category_names: Sequence[str],
    programme_codes: Sequence[str],
    modules: Sequence[str],
) -> dict[tuple[str, str], int]:
    """
    Fallback: count distinct suspended students per module from assessment detail
    when a suspended / enrollment flag column exists.
    """
    try:
        mart_cols = _mart_columns(conn, schema, TABLE_ASSESSMENT)
    except duckdb.CatalogException:
        return {}

    suspended_flag = _pick_first_mart_column(
        mart_cols,
        (
            "is_suspended",
            "user_is_suspended",
            "enrollment_suspended",
            "suspended",
        ),
    )
    user_status = _pick_first_mart_column(
        mart_cols, ("user_status", "enrollment_status", "student_status")
    )
    programme_col = _pick_first_mart_column(
        mart_cols, ("programme", "course_prefix", "program_code")
    )
    module_col = _pick_first_mart_column(
        mart_cols, ("course_shortname", "module_code")
    )
    user_col = _pick_first_mart_column(
        mart_cols, ("user_id", "student_no", "user_idnumber", "username")
    )
    if not programme_col or not module_col or not user_col:
        return {}
    if not suspended_flag and not user_status:
        return {}

    if suspended_flag:
        suspended_pred = (
            f'LOWER(TRIM(CAST("{suspended_flag}" AS VARCHAR))) IN '
            f"('true', 't', '1', 'yes', 'y')"
        )
    else:
        suspended_pred = (
            f'LOWER(TRIM(CAST("{user_status}" AS VARCHAR))) = \'suspended\''
        )

    # Reuse category/programme/module filters; do not apply assessment/status filters
    # so suspended enrollment counts reflect the module population.
    where_parts = [suspended_pred]
    params: list[Any] = []
    category_col = _pick_first_mart_column(mart_cols, ("category_name",))
    if category_names and category_col:
        where_parts.append(
            _in_clause(
                f'TRIM(CAST("{category_col}" AS VARCHAR))',
                [c.strip() for c in category_names],
                params,
            )
        )
    if programme_codes:
        where_parts.append(
            _in_clause(
                f'UPPER(TRIM(CAST("{programme_col}" AS VARCHAR)))',
                [c.strip().upper() for c in programme_codes],
                params,
            )
        )
    if modules:
        where_parts.append(
            _in_clause(
                f'UPPER(TRIM(CAST("{module_col}" AS VARCHAR)))',
                [m.strip().upper() for m in modules],
                params,
            )
        )

    relation = qualified_relation(schema, TABLE_ASSESSMENT)
    query = f"""
        SELECT
            UPPER(TRIM(CAST("{programme_col}" AS VARCHAR))) AS programme,
            UPPER(TRIM(CAST("{module_col}" AS VARCHAR))) AS module_code,
            COUNT(DISTINCT "{user_col}") AS suspended_students
        FROM {relation}
        WHERE {' AND '.join(where_parts)}
        GROUP BY 1, 2
    """
    try:
        rows = conn.execute(query, params).fetchall()
    except Exception:
        return {}

    out: dict[tuple[str, str], int] = {}
    for programme, module_code, count in rows:
        key = (str(programme or "").upper(), str(module_code or "").upper())
        if key[0] and key[1]:
            out[key] = int(count or 0)
    return out


def write_activity_completion_summary(
    ws: Worksheet,
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    *,
    category_names: Sequence[str],
    programme_codes: Sequence[str],
    modules: Sequence[str],
    assessment_types: Sequence[str],
    assessments: Sequence[str],
    statuses: Sequence[str],
) -> int:
    write_headers(ws, SUMMARY_HEADERS)
    widths = [max(12, min(len(h) + 2, MAX_COLUMN_WIDTH)) for h in SUMMARY_HEADERS]
    sample_remaining = [COLUMN_WIDTH_SAMPLE_ROWS]

    try:
        mart_cols = _mart_columns(conn, schema, TABLE_ASSESSMENT)
    except duckdb.CatalogException:
        finish_sheet(ws, SUMMARY_HEADERS, 0, widths)
        return 0

    category_col = _pick_first_mart_column(mart_cols, ("category_name",))
    programme_col = _pick_first_mart_column(
        mart_cols, ("programme", "course_prefix", "program_code")
    )
    module_code_col = _pick_first_mart_column(
        mart_cols, ("course_shortname", "module_code")
    )
    module_name_col = _pick_first_mart_column(mart_cols, ("course_fullname", "module"))
    assessment_col = _pick_first_mart_column(
        mart_cols, ("assessment", "assessment_name")
    )
    assessment_type_col = _pick_first_mart_column(mart_cols, ("assessment_type",))
    if not programme_col or not module_code_col or not assessment_col:
        finish_sheet(ws, SUMMARY_HEADERS, 0, widths)
        return 0

    select_parts = [
        (f'MAX("{category_col}") AS category_name' if category_col else "'' AS category_name"),
        f'"{programme_col}" AS programme',
        f'"{module_code_col}" AS module_code',
        (
            f'MAX("{module_name_col}") AS module_name'
            if module_name_col
            else "'' AS module_name"
        ),
        f'"{assessment_col}" AS assessment_name',
        (
            f'MAX("{assessment_type_col}") AS assessment_type'
            if assessment_type_col
            else "'' AS assessment_type"
        ),
        "COUNT(*) AS total_submissions",
    ]
    group_parts = [f'"{programme_col}"', f'"{module_code_col}"', f'"{assessment_col}"']

    built = _build_activity_filter_sql(
        conn,
        schema,
        category_names=category_names,
        programme_codes=programme_codes,
        modules=modules,
        assessment_types=assessment_types,
        assessments=assessments,
        statuses=(),
        mark_statuses=statuses,
        select_sql=", ".join(select_parts),
        group_by_sql=", ".join(group_parts),
        order_columns=["programme", "course_shortname", "assessment_name"],
    )
    if not built:
        finish_sheet(ws, SUMMARY_HEADERS, 0, widths)
        return 0

    query, params = built
    module_stats = _module_stats_lookup(
        conn,
        schema,
        category_names=category_names,
        programme_codes=programme_codes,
        modules=modules,
    )
    suspended_fallback = _suspended_by_module_from_detail(
        conn,
        schema,
        category_names=category_names,
        programme_codes=programme_codes,
        modules=modules,
    )

    result = conn.execute(query, params)
    columns = [str(desc[0]).lower() for desc in result.description]
    count = 0
    while True:
        batch = result.fetchmany(FETCH_CHUNK_SIZE)
        if not batch:
            break
        for tup in batch:
            row = normalize_row(dict(zip(columns, tup)))
            programme = str(pick(row, "programme") or "").strip().upper()
            module_code = str(pick(row, "module_code") or "").strip().upper()
            stats = module_stats.get((programme, module_code), {})
            students = pick(stats, "students", "total_students")
            suspended = pick(stats, "suspended_students", "suspended")
            if suspended == "" or suspended is None:
                suspended = suspended_fallback.get((programme, module_code), 0)
            module_name = pick(row, "module_name") or pick(stats, "module_name", "module")
            category = pick(row, "category_name") or pick(stats, "category_name")
            values = [
                format_cell(category),
                format_cell(pick(row, "programme")),
                format_cell(pick(row, "module_code")),
                format_cell(module_name),
                format_cell(pick(row, "assessment_name")),
                format_cell(pick(row, "assessment_type")),
                format_cell(students if students != "" else 0),
                format_cell(suspended if suspended != "" else 0),
                format_cell(pick(row, "total_submissions")),
            ]
            _update_col_widths(widths, values, sample_remaining)
            append_data_row(ws, values)
            count += 1

    finish_sheet(ws, SUMMARY_HEADERS, count, widths)
    return count


def write_submission_details(
    ws: Worksheet,
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    *,
    category_names: Sequence[str],
    programme_codes: Sequence[str],
    modules: Sequence[str],
    assessment_types: Sequence[str],
    assessments: Sequence[str],
    statuses: Sequence[str],
) -> int:
    write_headers(ws, DETAIL_HEADERS)
    widths = [max(12, min(len(h) + 2, MAX_COLUMN_WIDTH)) for h in DETAIL_HEADERS]
    sample_remaining = [COLUMN_WIDTH_SAMPLE_ROWS]
    count = 0

    for row in iter_filtered_assessment_rows(
        conn,
        schema,
        category_names=category_names,
        programme_codes=programme_codes,
        modules=modules,
        assessment_types=assessment_types,
        assessments=assessments,
        statuses=(),
        mark_statuses=statuses,
        order_columns=[
            "category_name",
            "programme",
            "course_prefix",
            "course_shortname",
            "assessment_name",
            "student_no",
        ],
    ):
        submitted = pick(
            row,
            "grade_submitted_at",
            "last_attempt_at",
            "submitted_at",
            "graded_at",
        )
        status_val = effective_completion_status(row)
        if not submitted and status_val not in SUBMISSION_STATUSES:
            submitted = ""

        values = [
            format_cell(pick(row, "category_name")),
            format_cell(pick(row, "programme", "course_prefix", "program_code")),
            format_cell(pick(row, "student_no", "user_username")),
            format_cell(pick(row, "user_fullname")),
            format_cell(pick(row, "user_email")),
            format_cell(pick(row, "course_shortname", "module_code")),
            format_cell(pick(row, "course_fullname", "module")),
            format_cell(pick(row, "assessment", "assessment_name")),
            format_cell(pick(row, "assessment_type")),
            format_cell(pick(row, "due_at", "effective_deadline_at")),
            format_cell(submitted),
            format_cell(status_val),
            format_cell(pick(row, "mark_status")),
            format_cell(pick(row, "grade_raw")),
            format_cell(pick(row, "max_grade")),
            *[format_cell(pick(row, *aliases)) for _, aliases in NOTE_FIELD_MAP],
        ]
        _update_col_widths(widths, values, sample_remaining)
        append_data_row(ws, values)
        count += 1

    finish_sheet(ws, DETAIL_HEADERS, count, widths)
    return count


def build_workbook(
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    output_dir: Path,
    *,
    category_names: Sequence[str],
    programme_codes: Sequence[str],
    modules: Sequence[str],
    assessment_types: Sequence[str],
    assessments: Sequence[str],
    statuses: Sequence[str],
) -> Path:
    wb = Workbook(write_only=True)

    ws_summary = wb.create_sheet(title="Activity Completion Summary"[:31])
    write_activity_completion_summary(
        ws_summary,
        conn,
        schema,
        category_names=category_names,
        programme_codes=programme_codes,
        modules=modules,
        assessment_types=assessment_types,
        assessments=assessments,
        statuses=statuses,
    )
    gc.collect()

    ws_detail = wb.create_sheet(title="Submission Details"[:31])
    write_submission_details(
        ws_detail,
        conn,
        schema,
        category_names=category_names,
        programme_codes=programme_codes,
        modules=modules,
        assessment_types=assessment_types,
        assessments=assessments,
        statuses=statuses,
    )
    gc.collect()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if len(programme_codes) == 1:
        file_code = programme_codes[0]
    elif programme_codes:
        joined = "_".join(programme_codes)
        file_code = joined if len(joined) <= 48 else f"batch_{len(programme_codes)}prog"
    else:
        file_code = "all_programmes"
    safe_code = file_code.replace(" ", "_")
    out_path = output_dir / f"activity_completion_{safe_code}_{timestamp}.xlsx"
    wb.save(out_path)
    return out_path.resolve()


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Export Activity Completion Excel from warehouse marts. "
            "Omit a filter (or pass *) for select-all on that dimension."
        )
    )
    parser.add_argument(
        "--category-name",
        action="append",
        dest="category_names",
        default=None,
        help="Category / intake name (repeatable; omit = all)",
    )
    parser.add_argument(
        "--programme-code",
        action="append",
        dest="programme_codes",
        default=None,
        help="Programme code (repeatable; omit = all)",
    )
    parser.add_argument(
        "--module",
        action="append",
        dest="modules",
        default=None,
        help="Module code / course shortname (repeatable; omit = all)",
    )
    parser.add_argument(
        "--assessment-type",
        action="append",
        dest="assessment_types",
        default=None,
        help="Assessment type (repeatable; omit = all)",
    )
    parser.add_argument(
        "--assessment",
        action="append",
        dest="assessments",
        default=None,
        help="Assessment name (repeatable; omit = all)",
    )
    parser.add_argument(
        "--status",
        action="append",
        dest="statuses",
        default=None,
        help=(
            "Mark status filter for activity completion (repeatable): "
            f"{', '.join(MARK_STATUSES)}. Omit = all."
        ),
    )
    parser.add_argument(
        "--warehouse-schema",
        default=None,
        help="Schema for gradebook marts (default: moodle_processed)",
    )
    parser.add_argument(
        "--output-dir",
        default=".",
        help="Directory for the generated workbook",
    )
    args = parser.parse_args()

    category_names = _normalize_filter_values(args.category_names)
    programme_codes = [
        c.strip().upper()
        for c in _normalize_filter_values(args.programme_codes)
    ]
    modules = [m.strip().upper() for m in _normalize_filter_values(args.modules)]
    assessment_types = [
        t.strip().upper() for t in _normalize_filter_values(args.assessment_types)
    ]
    assessments = [
        a.strip().upper() for a in _normalize_filter_values(args.assessments)
    ]
    statuses = _normalize_mark_statuses(args.statuses)

    schema = args.warehouse_schema or gradebook_schema() or DEFAULT_SCHEMA
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    conn = connect_motherduck()
    try:
        out_path = build_workbook(
            conn,
            schema,
            output_dir,
            category_names=category_names,
            programme_codes=programme_codes,
            modules=modules,
            assessment_types=assessment_types,
            assessments=assessments,
            statuses=statuses,
        )
    except Exception:
        import traceback

        traceback.print_exc()
        raise
    finally:
        conn.close()

    print(out_path)


if __name__ == "__main__":
    main()
