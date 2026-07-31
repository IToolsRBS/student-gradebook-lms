"""
Build Inactivity Report workbook from warehouse gradebook marts.

Filters (CLI; empty / omitted = select all for category/programme):
  category, programme, inactivity period (7 / 14 / 30 / never).

Sheets:
  1. Inactivity Summary — one row per programme
       - total students registered
       - inactive students for the chosen period
       - students with no submissions/late submissions (all past-due assessments missed)
  2. Inactive Students — students with no Moodle access in the chosen period
  3. Never Submitted — no submissions/late; all past-due assessments are missed

Uses gradebook_student_summary (access + total_submitted) and, when needed,
gradebook_student_assessment_detail (joined via user_idnumber or user_username).

Prints the absolute output path as the last stdout line (for future frontend wiring).
"""

from __future__ import annotations

import argparse
import gc
from datetime import datetime
from pathlib import Path
from typing import Any, Iterator, Sequence

import duckdb
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from motherduck_client import (
    connect_motherduck,
    gradebook_schema,
    qualified_relation,
)
from populate_activity_completion import (
    _in_clause,
    _normalize_filter_values,
)
from populate_gradebook_from_warehouse import (
    COLUMN_WIDTH_SAMPLE_ROWS,
    DEFAULT_SCHEMA,
    FETCH_CHUNK_SIZE,
    MAX_COLUMN_WIDTH,
    TABLE_ASSESSMENT,
    TABLE_STUDENT,
    _mart_columns,
    _pick_first_mart_column,
    _update_col_widths,
    append_data_row,
    finish_sheet,
    format_cell,
    normalize_row,
    pick,
    write_headers,
)

INACTIVITY_PERIODS: dict[str, int | None] = {
    "7": 7,
    "14": 14,
    "30": 30,
    "never": None,
}

INACTIVE_STUDENT_HEADERS: list[str] = [
    "Category",
    "Programme",
    "Student No",
    "Student",
    "Email",
    "Status",
    "Mark Status",
    "Modules",
    "Last Moodle Access",
    "Days Since Access",
]

NEVER_SUBMITTED_HEADERS: list[str] = [
    "Category",
    "Programme",
    "Student No",
    "Student",
    "Email",
    "Status",
    "Mark Status",
    "Modules",
    "Last Moodle Access",
    "Days Since Access",
    "Total Submissions Made",
]


def _normalize_inactivity_period(raw: str | None) -> str:
    key = str(raw or "7").strip().lower().replace(" ", "").replace("_", "")
    aliases = {
        "7": "7",
        "7days": "7",
        "7d": "7",
        "14": "14",
        "14days": "14",
        "14d": "14",
        "30": "30",
        "30days": "30",
        "30d": "30",
        "never": "never",
        "neveraccessed": "never",
        "noaccess": "never",
    }
    period = aliases.get(key)
    if not period:
        raise SystemExit(
            f"Invalid --inactivity-period '{raw}'. "
            f"Allowed: 7, 14, 30, never"
        )
    return period


def _period_label(period: str) -> str:
    if period == "never":
        return "Never accessed"
    return f"{period} days"


def _student_dimension_where(
    mart_cols: dict[str, str],
    *,
    category_names: Sequence[str],
    programme_codes: Sequence[str],
) -> tuple[list[str], list[Any]]:
    where_parts: list[str] = []
    params: list[Any] = []

    category_col = _pick_first_mart_column(mart_cols, ("category_name",))
    programme_col = _pick_first_mart_column(
        mart_cols, ("programme", "program_code", "course_prefix")
    )

    if category_names and category_col:
        where_parts.append(
            _in_clause(
                f'TRIM(CAST("{category_col}" AS VARCHAR))',
                [c.strip() for c in category_names],
                params,
            )
        )
    if programme_codes and programme_col:
        where_parts.append(
            _in_clause(
                f'UPPER(TRIM(CAST("{programme_col}" AS VARCHAR)))',
                [c.strip().upper() for c in programme_codes],
                params,
            )
        )
    return where_parts, params


def _inactivity_predicate_sql(
    mart_cols: dict[str, str],
    period: str,
) -> str | None:
    """
    SQL boolean for students inactive under the chosen period.

    - 7/14/30: days_since_access >= N, or never accessed (null access)
    - never: last access is null / blank
    """
    days_col = _pick_first_mart_column(
        mart_cols, ("days_since_access", "days_since_last_access")
    )
    access_col = _pick_first_mart_column(
        mart_cols,
        ("last_moodle_access", "last_access", "lastaccess", "last_login"),
    )
    if not days_col and not access_col:
        return None

    never_parts: list[str] = []
    if access_col:
        never_parts.append(f'"{access_col}" IS NULL')
        never_parts.append(
            f"TRIM(CAST(\"{access_col}\" AS VARCHAR)) IN ('', 'nan', 'none', 'null')"
        )
    if days_col:
        never_parts.append(f'"{days_col}" IS NULL')
    never_pred = " OR ".join(never_parts) if never_parts else "false"

    if period == "never":
        return f"({never_pred})"

    days = INACTIVITY_PERIODS[period]
    assert days is not None
    if days_col:
        inactive_days = (
            f'(TRY_CAST("{days_col}" AS DOUBLE) IS NOT NULL '
            f'AND TRY_CAST("{days_col}" AS DOUBLE) >= {int(days)})'
        )
        return f"({inactive_days} OR ({never_pred}))"
    # Fallback: last access older than N days.
    assert access_col
    return (
        f"(({never_pred}) OR "
        f'(TRY_CAST("{access_col}" AS TIMESTAMP) IS NOT NULL AND '
        f'TRY_CAST("{access_col}" AS TIMESTAMP) <= '
        f"CURRENT_TIMESTAMP - INTERVAL '{int(days)}' DAY))"
    )


def _student_detail_join_parts(
    student_cols: dict[str, str],
    detail_cols: dict[str, str],
) -> tuple[list[str], str, str] | None:
    """Join keys between student summary (s) and assessment detail (d)."""
    student_key = _pick_first_mart_column(
        student_cols, ("student_no", "user_idnumber", "username", "student_id")
    )
    # Prefer user_idnumber when populated; many Moodle rows only store the
    # student number on user_username (user_idnumber null).
    detail_idnumber = _pick_first_mart_column(detail_cols, ("user_idnumber", "student_no"))
    detail_username = _pick_first_mart_column(
        detail_cols, ("user_username", "username")
    )
    if not student_key or (not detail_idnumber and not detail_username):
        return None

    student_expr = f'UPPER(TRIM(CAST(s."{student_key}" AS VARCHAR)))'
    id_parts: list[str] = []
    if detail_idnumber:
        id_parts.append(
            f'UPPER(TRIM(CAST(d."{detail_idnumber}" AS VARCHAR))) = {student_expr}'
        )
    if detail_username:
        id_parts.append(
            f'UPPER(TRIM(CAST(d."{detail_username}" AS VARCHAR))) = {student_expr}'
        )
    join_parts = [
        id_parts[0] if len(id_parts) == 1 else "(" + " OR ".join(id_parts) + ")"
    ]

    stu_prog = _pick_first_mart_column(
        student_cols, ("programme", "program_code", "course_prefix")
    )
    det_prog = _pick_first_mart_column(
        detail_cols, ("programme", "course_prefix", "program_code")
    )
    stu_cat = _pick_first_mart_column(student_cols, ("category_name",))
    det_cat = _pick_first_mart_column(detail_cols, ("category_name",))
    if stu_prog and det_prog:
        join_parts.append(
            f'UPPER(TRIM(CAST(d."{det_prog}" AS VARCHAR))) = '
            f'UPPER(TRIM(CAST(s."{stu_prog}" AS VARCHAR)))'
        )
    if stu_cat and det_cat:
        join_parts.append(
            f'TRIM(CAST(d."{det_cat}" AS VARCHAR)) = '
            f'TRIM(CAST(s."{stu_cat}" AS VARCHAR))'
        )
    detail_key = detail_idnumber or detail_username or ""
    return join_parts, student_key, detail_key


def _submission_evidence_sql(detail_cols: dict[str, str], alias: str = "d") -> str:
    """
    SQL predicate: this assessment detail row shows a real submission/grade.
    Aligns with activity-completion effective status (submitted / graded / late).
    """
    prefix = f"{alias}."
    status_col = _pick_first_mart_column(detail_cols, ("status",))
    is_submitted_col = _pick_first_mart_column(detail_cols, ("is_submitted",))
    has_attempt_col = _pick_first_mart_column(detail_cols, ("has_attempt",))
    attempt_count_col = _pick_first_mart_column(detail_cols, ("attempt_count",))
    grade_col = _pick_first_mart_column(
        detail_cols, ("grade_raw", "final_grade", "grade")
    )
    graded_at_col = _pick_first_mart_column(detail_cols, ("graded_at",))
    submitted_at_col = _pick_first_mart_column(
        detail_cols,
        ("grade_submitted_at", "last_attempt_at", "submitted_at"),
    )

    parts: list[str] = []
    if status_col:
        parts.append(
            "LOWER(REPLACE(REPLACE(TRIM(CAST("
            f'{prefix}"{status_col}" AS VARCHAR)), \' \', \'_\'), \'-\', \'_\')) '
            "IN ('submitted', 'graded', 'submitted_late')"
        )
    if is_submitted_col:
        col = f'{prefix}"{is_submitted_col}"'
        parts.append(
            f"(COALESCE(TRY_CAST({col} AS BOOLEAN), false) = true "
            f"OR LOWER(TRIM(CAST({col} AS VARCHAR))) IN ('true', 't', '1', 'yes', 'y'))"
        )
    if has_attempt_col:
        col = f'{prefix}"{has_attempt_col}"'
        parts.append(
            f"(COALESCE(TRY_CAST({col} AS BOOLEAN), false) = true "
            f"OR LOWER(TRIM(CAST({col} AS VARCHAR))) IN ('true', 't', '1', 'yes', 'y'))"
        )
    if attempt_count_col:
        col = f'{prefix}"{attempt_count_col}"'
        parts.append(
            f"(TRY_CAST({col} AS BIGINT) IS NOT NULL AND TRY_CAST({col} AS BIGINT) > 0)"
        )
    if grade_col:
        col = f'{prefix}"{grade_col}"'
        parts.append(
            f"({col} IS NOT NULL AND "
            f"TRIM(CAST({col} AS VARCHAR)) NOT IN ('', 'nan', 'none', 'null'))"
        )
    if graded_at_col:
        parts.append(f'{prefix}"{graded_at_col}" IS NOT NULL')
    if submitted_at_col:
        parts.append(f'{prefix}"{submitted_at_col}" IS NOT NULL')

    if not parts:
        return "false"
    return "(" + " OR ".join(parts) + ")"


def _past_due_sql(detail_cols: dict[str, str], alias: str = "d") -> str:
    """SQL predicate: assessment is past its due / effective deadline."""
    prefix = f"{alias}."
    status_col = _pick_first_mart_column(detail_cols, ("status",))
    due_col = _pick_first_mart_column(
        detail_cols, ("due_at", "effective_deadline_at", "due_date")
    )
    days_overdue_col = _pick_first_mart_column(detail_cols, ("days_overdue",))

    parts: list[str] = []
    if status_col:
        # Warehouse already classifies these as past-due non-submissions.
        parts.append(
            "LOWER(REPLACE(REPLACE(TRIM(CAST("
            f'{prefix}"{status_col}" AS VARCHAR)), \' \', \'_\'), \'-\', \'_\')) '
            "= 'missed'"
        )
    if due_col:
        parts.append(
            f'(TRY_CAST({prefix}"{due_col}" AS TIMESTAMP) IS NOT NULL AND '
            f'TRY_CAST({prefix}"{due_col}" AS TIMESTAMP) < CURRENT_TIMESTAMP)'
        )
    if days_overdue_col:
        parts.append(
            f'(TRY_CAST({prefix}"{days_overdue_col}" AS DOUBLE) IS NOT NULL AND '
            f'TRY_CAST({prefix}"{days_overdue_col}" AS DOUBLE) > 0)'
        )
    if not parts:
        return "false"
    return "(" + " OR ".join(parts) + ")"


def _status_is_missed_sql(detail_cols: dict[str, str], alias: str = "d") -> str:
    status_col = _pick_first_mart_column(detail_cols, ("status",))
    if not status_col:
        return "false"
    return (
        "LOWER(REPLACE(REPLACE(TRIM(CAST("
        f'{alias}."{status_col}" AS VARCHAR)), \' \', \'_\'), \'-\', \'_\')) '
        "= 'missed'"
    )


def _never_submitted_predicate(
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    student_cols: dict[str, str],
) -> tuple[str, list[Any]] | None:
    """
    True when the student has no submissions / late submissions in scope, and
    every past-due assessment is status 'missed'.

    Future assessments (not_due / open before deadline) are allowed.
    """
    try:
        detail_cols = _mart_columns(conn, schema, TABLE_ASSESSMENT)
    except duckdb.CatalogException:
        return None

    join = _student_detail_join_parts(student_cols, detail_cols)
    if not join:
        return None
    join_parts, _student_key, _detail_key = join
    detail_rel = qualified_relation(schema, TABLE_ASSESSMENT)
    join_sql = " AND ".join(join_parts)
    past_due = _past_due_sql(detail_cols, alias="d")
    is_missed = _status_is_missed_sql(detail_cols, alias="d")
    submitted = _submission_evidence_sql(detail_cols, alias="d")

    # Optional summary guards when present (total submitted + late counts).
    summary_guards: list[str] = []
    submitted_col = _pick_first_mart_column(
        student_cols,
        (
            "total_submitted",
            "submitted_assessments",
            "total_submissions",
            "submitted_count",
        ),
    )
    late_col = _pick_first_mart_column(
        student_cols, ("late_submissions", "late", "late_count")
    )
    if submitted_col:
        summary_guards.append(
            f'COALESCE(TRY_CAST(s."{submitted_col}" AS DOUBLE), 0) <= 0'
        )
    if late_col:
        summary_guards.append(
            f'COALESCE(TRY_CAST(s."{late_col}" AS DOUBLE), 0) <= 0'
        )
    summary_clause = (
        " AND ".join(summary_guards) if summary_guards else "true"
    )

    sql = f"""
        ({summary_clause})
        AND EXISTS (
            SELECT 1
            FROM {detail_rel} AS d
            WHERE {join_sql}
              AND {past_due}
        )
        AND NOT EXISTS (
            SELECT 1
            FROM {detail_rel} AS d
            WHERE {join_sql}
              AND {submitted}
        )
        AND NOT EXISTS (
            SELECT 1
            FROM {detail_rel} AS d
            WHERE {join_sql}
              AND {past_due}
              AND NOT ({is_missed})
        )
    """
    return sql.strip(), []


def _build_student_base_sql(
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    *,
    category_names: Sequence[str],
    programme_codes: Sequence[str],
    extra_where: str | None = None,
    extra_params: Sequence[Any] | None = None,
    order_columns: Sequence[str] | None = None,
) -> tuple[str, list[Any], dict[str, str]] | None:
    relation = qualified_relation(schema, TABLE_STUDENT)
    try:
        mart_cols = _mart_columns(conn, schema, TABLE_STUDENT)
    except duckdb.CatalogException:
        return None

    where_parts, params = _student_dimension_where(
        mart_cols,
        category_names=category_names,
        programme_codes=programme_codes,
    )
    if extra_where:
        where_parts.append(f"({extra_where})")
        params.extend(list(extra_params or []))

    base_where = " AND ".join(where_parts) if where_parts else "1=1"
    query = f'SELECT * FROM {relation} AS s WHERE {base_where}'
    if order_columns:
        present = [col for col in order_columns if col.lower() in mart_cols]
        if present:
            order = ", ".join(f's."{mart_cols[col.lower()]}"' for col in present)
            query += f" ORDER BY {order}"
        else:
            query += " ORDER BY 1"
    else:
        query += " ORDER BY 1"
    return query, params, mart_cols


def iter_student_rows(
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    *,
    category_names: Sequence[str],
    programme_codes: Sequence[str],
    extra_where: str | None = None,
    extra_params: Sequence[Any] | None = None,
    order_columns: Sequence[str] | None = None,
    chunk_size: int = FETCH_CHUNK_SIZE,
) -> Iterator[dict[str, Any]]:
    built = _build_student_base_sql(
        conn,
        schema,
        category_names=category_names,
        programme_codes=programme_codes,
        extra_where=extra_where,
        extra_params=extra_params,
        order_columns=order_columns,
    )
    if not built:
        return
    query, params, _ = built
    result = conn.execute(query, params)
    columns = [str(desc[0]) for desc in result.description]
    while True:
        batch = result.fetchmany(chunk_size)
        if not batch:
            break
        for tup in batch:
            yield normalize_row(dict(zip(columns, tup)))


def _write_inactive_student_rows(
    ws: Worksheet,
    rows: Iterator[dict[str, Any]],
) -> int:
    write_headers(ws, INACTIVE_STUDENT_HEADERS)
    widths = [
        max(12, min(len(h) + 2, MAX_COLUMN_WIDTH)) for h in INACTIVE_STUDENT_HEADERS
    ]
    sample_remaining = [COLUMN_WIDTH_SAMPLE_ROWS]
    count = 0
    for row in rows:
        values = [
            format_cell(pick(row, "category_name")),
            format_cell(pick(row, "programme", "program_code")),
            format_cell(pick(row, "student_no")),
            format_cell(pick(row, "student", "user_fullname")),
            format_cell(pick(row, "email", "user_email")),
            format_cell(pick(row, "status")),
            format_cell(pick(row, "mark_status")),
            format_cell(pick(row, "total_modules", "modules")),
            format_cell(pick(row, "last_moodle_access", "last_access")),
            format_cell(pick(row, "days_since_access", "days_since_last_access")),
        ]
        _update_col_widths(widths, values, sample_remaining)
        append_data_row(ws, values)
        count += 1
    finish_sheet(ws, len(INACTIVE_STUDENT_HEADERS), count, widths)
    return count


def _write_never_submitted_rows(
    ws: Worksheet,
    rows: Iterator[dict[str, Any]],
) -> int:
    write_headers(ws, NEVER_SUBMITTED_HEADERS)
    widths = [
        max(12, min(len(h) + 2, MAX_COLUMN_WIDTH)) for h in NEVER_SUBMITTED_HEADERS
    ]
    sample_remaining = [COLUMN_WIDTH_SAMPLE_ROWS]
    count = 0
    for row in rows:
        values = [
            format_cell(pick(row, "category_name")),
            format_cell(pick(row, "programme", "program_code")),
            format_cell(pick(row, "student_no")),
            format_cell(pick(row, "student", "user_fullname")),
            format_cell(pick(row, "email", "user_email")),
            format_cell(pick(row, "status")),
            format_cell(pick(row, "mark_status")),
            format_cell(pick(row, "total_modules", "modules")),
            format_cell(pick(row, "last_moodle_access", "last_access")),
            format_cell(pick(row, "days_since_access", "days_since_last_access")),
            format_cell(0),  # Total Submissions Made — all missed
        ]
        _update_col_widths(widths, values, sample_remaining)
        append_data_row(ws, values)
        count += 1
    finish_sheet(ws, len(NEVER_SUBMITTED_HEADERS), count, widths)
    return count


def write_inactivity_summary(
    ws: Worksheet,
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    *,
    category_names: Sequence[str],
    programme_codes: Sequence[str],
    period: str,
) -> int:
    inactive_header = f"Inactive Students ({_period_label(period)})"
    headers = [
        "Category",
        "Programme",
        "Total Students Registered",
        inactive_header,
        "Never Submitted Assessment",
    ]
    write_headers(ws, headers)
    widths = [max(12, min(len(h) + 2, MAX_COLUMN_WIDTH)) for h in headers]
    sample_remaining = [COLUMN_WIDTH_SAMPLE_ROWS]

    try:
        mart_cols = _mart_columns(conn, schema, TABLE_STUDENT)
    except duckdb.CatalogException:
        finish_sheet(ws, len(headers), 0, widths)
        return 0

    inactive_pred = _inactivity_predicate_sql(mart_cols, period)
    never_sub = _never_submitted_predicate(conn, schema, mart_cols)
    if not inactive_pred:
        finish_sheet(ws, len(headers), 0, widths)
        return 0

    programme_col = _pick_first_mart_column(
        mart_cols, ("programme", "program_code", "course_prefix")
    )
    category_col = _pick_first_mart_column(mart_cols, ("category_name",))
    if not programme_col:
        finish_sheet(ws, len(headers), 0, widths)
        return 0

    where_parts, params = _student_dimension_where(
        mart_cols,
        category_names=category_names,
        programme_codes=programme_codes,
    )
    base_where = " AND ".join(where_parts) if where_parts else "1=1"
    relation = qualified_relation(schema, TABLE_STUDENT)

    never_expr = "0"
    if never_sub:
        never_sql, never_params = never_sub
        never_expr = f"SUM(CASE WHEN ({never_sql}) THEN 1 ELSE 0 END)"
        params = list(params) + list(never_params)

    select_parts = [
        (
            f'MAX(s."{category_col}") AS category_name'
            if category_col
            else "'' AS category_name"
        ),
        f's."{programme_col}" AS programme',
        "COUNT(*) AS total_students_registered",
        f"SUM(CASE WHEN ({inactive_pred}) THEN 1 ELSE 0 END) AS inactive_students",
        f"{never_expr} AS never_submitted",
    ]
    query = f"""
        SELECT {', '.join(select_parts)}
        FROM {relation} AS s
        WHERE {base_where}
        GROUP BY s."{programme_col}"
        ORDER BY s."{programme_col}"
    """

    try:
        result = conn.execute(query, params)
    except Exception:
        # Retry summary without never-submitted if detail join fails.
        select_parts[-1] = "0 AS never_submitted"
        query = f"""
            SELECT {', '.join(select_parts)}
            FROM {relation} AS s
            WHERE {base_where}
            GROUP BY s."{programme_col}"
            ORDER BY s."{programme_col}"
        """
        dim_params = _student_dimension_where(
            mart_cols,
            category_names=category_names,
            programme_codes=programme_codes,
        )[1]
        result = conn.execute(query, dim_params)

    columns = [str(desc[0]).lower() for desc in result.description]
    count = 0
    while True:
        batch = result.fetchmany(FETCH_CHUNK_SIZE)
        if not batch:
            break
        for tup in batch:
            row = normalize_row(dict(zip(columns, tup)))
            values = [
                format_cell(pick(row, "category_name")),
                format_cell(pick(row, "programme")),
                format_cell(pick(row, "total_students_registered")),
                format_cell(pick(row, "inactive_students")),
                format_cell(pick(row, "never_submitted")),
            ]
            _update_col_widths(widths, values, sample_remaining)
            append_data_row(ws, values)
            count += 1

    finish_sheet(ws, len(headers), count, widths)
    return count


def write_inactive_students(
    ws: Worksheet,
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    *,
    category_names: Sequence[str],
    programme_codes: Sequence[str],
    period: str,
) -> int:
    try:
        mart_cols = _mart_columns(conn, schema, TABLE_STUDENT)
    except duckdb.CatalogException:
        write_headers(ws, INACTIVE_STUDENT_HEADERS)
        finish_sheet(ws, len(INACTIVE_STUDENT_HEADERS), 0)
        return 0

    inactive_pred = _inactivity_predicate_sql(mart_cols, period)
    if not inactive_pred:
        write_headers(ws, INACTIVE_STUDENT_HEADERS)
        finish_sheet(ws, len(INACTIVE_STUDENT_HEADERS), 0)
        return 0

    return _write_inactive_student_rows(
        ws,
        iter_student_rows(
            conn,
            schema,
            category_names=category_names,
            programme_codes=programme_codes,
            extra_where=inactive_pred,
            order_columns=["programme", "days_since_access", "student_no"],
        ),
    )


def write_never_submitted_students(
    ws: Worksheet,
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    *,
    category_names: Sequence[str],
    programme_codes: Sequence[str],
) -> int:
    try:
        mart_cols = _mart_columns(conn, schema, TABLE_STUDENT)
    except duckdb.CatalogException:
        write_headers(ws, NEVER_SUBMITTED_HEADERS)
        finish_sheet(ws, len(NEVER_SUBMITTED_HEADERS), 0)
        return 0

    never_sub = _never_submitted_predicate(conn, schema, mart_cols)
    if not never_sub:
        write_headers(ws, NEVER_SUBMITTED_HEADERS)
        finish_sheet(ws, len(NEVER_SUBMITTED_HEADERS), 0)
        return 0

    never_sql, never_params = never_sub
    return _write_never_submitted_rows(
        ws,
        iter_student_rows(
            conn,
            schema,
            category_names=category_names,
            programme_codes=programme_codes,
            extra_where=never_sql,
            extra_params=never_params,
            order_columns=["programme", "student_no"],
        ),
    )


def build_workbook(
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    output_dir: Path,
    *,
    category_names: Sequence[str],
    programme_codes: Sequence[str],
    period: str,
) -> Path:
    wb = Workbook(write_only=True)

    ws_summary = wb.create_sheet(title="Inactivity Summary"[:31])
    write_inactivity_summary(
        ws_summary,
        conn,
        schema,
        category_names=category_names,
        programme_codes=programme_codes,
        period=period,
    )
    gc.collect()

    inactive_title = f"Inactive ({_period_label(period)})"[:31]
    ws_inactive = wb.create_sheet(title=inactive_title)
    write_inactive_students(
        ws_inactive,
        conn,
        schema,
        category_names=category_names,
        programme_codes=programme_codes,
        period=period,
    )
    gc.collect()

    ws_never = wb.create_sheet(title="Never Submitted"[:31])
    write_never_submitted_students(
        ws_never,
        conn,
        schema,
        category_names=category_names,
        programme_codes=programme_codes,
    )
    gc.collect()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if len(programme_codes) == 1:
        file_code = programme_codes[0]
    elif programme_codes:
        joined = "_".join(programme_codes)
        file_code = joined if len(joined) <= 48 else f"batch_{len(programme_codes)}prog"
    else:
        file_code = "all_programmes"
    safe_code = file_code.replace(" ", "_")
    out_path = (
        output_dir / f"inactivity_report_{period}d_{safe_code}_{timestamp}.xlsx"
        if period != "never"
        else output_dir / f"inactivity_report_never_{safe_code}_{timestamp}.xlsx"
    )
    wb.save(out_path)
    return out_path.resolve()


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Export Inactivity Report Excel from warehouse marts. "
            "Omit category/programme for select-all. "
            "Inactivity period: 7, 14, 30, or never."
        )
    )
    parser.add_argument(
        "--category-name",
        action="append",
        dest="category_names",
        default=None,
        help="Category / intake name (repeatable; omit = all)",
    )
    parser.add_argument(
        "--programme-code",
        action="append",
        dest="programme_codes",
        default=None,
        help="Programme code (repeatable; omit = all)",
    )
    parser.add_argument(
        "--inactivity-period",
        default="7",
        help="Inactivity window: 7, 14, 30, or never (default: 7)",
    )
    parser.add_argument(
        "--warehouse-schema",
        default=None,
        help="Schema for gradebook marts (default: moodle_processed)",
    )
    parser.add_argument(
        "--output-dir",
        default=".",
        help="Directory for the generated workbook",
    )
    args = parser.parse_args()

    category_names = _normalize_filter_values(args.category_names)
    programme_codes = [
        c.strip().upper() for c in _normalize_filter_values(args.programme_codes)
    ]
    period = _normalize_inactivity_period(args.inactivity_period)

    schema = args.warehouse_schema or gradebook_schema() or DEFAULT_SCHEMA
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    conn = connect_motherduck()
    try:
        out_path = build_workbook(
            conn,
            schema,
            output_dir,
            category_names=category_names,
            programme_codes=programme_codes,
            period=period,
        )
    except Exception:
        import traceback

        traceback.print_exc()
        raise
    finally:
        conn.close()

    print(out_path)


if __name__ == "__main__":
    main()
