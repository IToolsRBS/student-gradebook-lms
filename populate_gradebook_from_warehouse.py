"""
Build Gradebook export workbook from warehouse dbt marts (moodle_processed.*).

Reads pre-built tables; does not sync Moodle or run dbt.
Prints the absolute output path as the last stdout line for frontend/server.js.
"""

from __future__ import annotations

import argparse
import gc
import math
import warnings
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterator, Sequence

import duckdb
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

# Stream MotherDuck results in chunks to avoid pandas DataFrame spikes.
FETCH_CHUNK_SIZE = 2000
# Kept for callers that still sample widths (write-only uses heuristics instead).
COLUMN_WIDTH_SAMPLE_ROWS = 200
MIN_COLUMN_WIDTH = 10
DEFAULT_COLUMN_WIDTH = 12
MAX_COLUMN_WIDTH = 35
MAX_WRAP_COLUMN_WIDTH = 45

# Regent Business School export styling
_BRAND_MAROON = "632523"
_SUSPENDED_GREY = "D9D9D9"
_BORDER_GREY = "BFBFBF"
_FONT_NAME = "Calibri"
_FONT_SIZE = 11

from motherduck_client import (
    connect_motherduck,
    gradebook_schema,
    qualified_relation,
    read_env_value,
)
from warehouse_export_fallback import build_workbook_offering_fallback

DEFAULT_SCHEMA = "moodle_processed"

# Mart table names (within WAREHOUSE_GRADEBOOK_SCHEMA)
TABLE_PROGRAMME = "gradebook_programme_summary"
TABLE_STUDENT = "gradebook_student_summary"
TABLE_MODULE = "gradebook_module_summary"
TABLE_TRENDS = "gradebook_submission_trends"
TABLE_ASSESSMENT = "gradebook_student_assessment_detail"
TABLE_MISSED = "gradebook_missed_assessments"
TABLE_UPCOMING = "gradebook_upcoming_deadlines"
TABLE_COURSE_NOTES = "gradebook_course_notes"
COURSE_NOTES_SHEET_TITLE = "Course Notes"
COURSE_NOTES_COLUMNS: tuple[str, ...] = (
    "user_full_name",
    "username",
    "course_display_name",
    "content",
    "timestamp",
    "staff_id",
)

NOTE_FIELD_MAP: list[tuple[str, tuple[str, ...]]] = [
    ("Latest note", ("latest_note_content",)),
    ("Note timestamp", ("latest_note_timestamp",)),
    ("Note created by", ("latest_note_staff_id",)),
    ("Note course", ("latest_note_course_display_name",)),
]

# Column headers for mart sheets (used when workbooks need empty placeholder tabs).
MART_SHEET_HEADERS: dict[str, list[str]] = {
    "Submission Trends": [
        "Programme",
        "Module Code",
        "Module",
        "Assessment",
        "Assessment Type",
        "Due Date",
        "Total Students",
        "Submitted Count",
        "Missed Count",
        "Late Submissions",
    ],
    "Missed Assessments": [
        "Programme",
        "Student No",
        "Student",
        "Email",
        "Module Code",
        "Module",
        "Assessment",
        "Assessment Type",
        "Due Date",
        "Effective Deadline",
        "Days Overdue",
        "Status",
        "Mark Status",
        "Latest note",
        "Note timestamp",
        "Note created by",
        "Note course",
    ],
    "Upcoming Deadlines": [
        "Programme",
        "Module Code",
        "Module",
        "Assessment",
        "Assessment Type",
        "Effective Deadline",
        "Hours Until Due",
    ],
}

PROGRAMME_SUMMARY_HEADERS: list[str] = [
    "Programme",
    "Students",
    "Suspended Students",
    "Active Modules",
    "Submitted Assessments",
    "Missed Assessments",
    "Late Submissions",
    "Upcoming Deadlines (14 Days)",
    "Students With Missed Assessments",
]

# Per-mart programme/category filter columns (warehouse schema).
# All gradebook marts are scoped by category_name (intake) + offering code.
# All marts filter on programme (dropdown value from gradebook_module_summary).
MART_FILTER: dict[str, dict[str, Any]] = {
    TABLE_PROGRAMME: {
        "programme_cols": ("programme", "program_code"),
        "category_cols": ("category_name",),
    },
    TABLE_STUDENT: {
        "programme_cols": ("programme", "program_code"),
        "category_cols": ("category_name",),
    },
    TABLE_MODULE: {
        "programme_cols": ("programme", "program_code"),
        "category_cols": ("category_name",),
    },
    TABLE_TRENDS: {
        "programme_cols": ("programme", "program_code"),
        "category_cols": ("category_name",),
    },
    TABLE_ASSESSMENT: {
        # programme when present; else course_prefix (not program_code — that is canonical).
        "programme_cols": ("programme", "course_prefix"),
        "category_cols": ("category_name",),
    },
    TABLE_MISSED: {
        "programme_cols": ("programme", "program_code"),
        "category_cols": ("category_name",),
    },
    TABLE_UPCOMING: {
        "programme_cols": ("programme",),
        "category_cols": ("category_name",),
    },
}


def _is_missing(value: Any) -> bool:
    """True for None / NaN without importing pandas."""
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    # numpy / pandas scalars expose dtype + item()
    if hasattr(value, "dtype") and hasattr(value, "item"):
        try:
            if str(getattr(value, "dtype", "")).startswith("float"):
                return bool(math.isnan(float(value)))
        except (TypeError, ValueError):
            return True
    return False


def _mart_columns(
    conn: duckdb.DuckDBPyConnection, schema: str, table: str
) -> dict[str, str]:
    """Lowercase column name -> actual mart column name (no pandas)."""
    relation = qualified_relation(schema, table)
    described = conn.execute(f"DESCRIBE SELECT * FROM {relation} LIMIT 0").fetchall()
    return {str(row[0]).lower(): str(row[0]) for row in described}


def _pick_first_mart_column(
    mart_cols: dict[str, str], candidates: Sequence[str]
) -> str | None:
    for candidate in candidates:
        actual = mart_cols.get(str(candidate).lower())
        if actual:
            return actual
    return None


def normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    return {str(k).lower(): v for k, v in row.items()}


def pick(row: dict[str, Any], *keys: str) -> Any:
    """Return first non-empty value for any of the given column names."""
    for key in keys:
        value = row.get(key.lower())
        if _is_missing(value):
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return value
    return ""


def _format_sequence(value: Any) -> str:
    if hasattr(value, "tolist") and not isinstance(value, (str, bytes)):
        try:
            value = value.tolist()
        except (TypeError, ValueError):
            pass
    if not isinstance(value, (list, tuple)):
        return str(value)
    parts: list[str] = []
    for item in value:
        if _is_missing(item):
            continue
        parts.append(str(item))
    return ", ".join(parts)


def _excel_safe_datetime(value: datetime) -> datetime:
    """Excel/openpyxl reject timezone-aware datetimes — strip tzinfo."""
    if getattr(value, "tzinfo", None) is not None:
        try:
            return value.replace(tzinfo=None)
        except (TypeError, ValueError, OSError):
            # Fall back to naive wall-clock components if replace fails.
            return datetime(
                value.year,
                value.month,
                value.day,
                value.hour,
                value.minute,
                value.second,
                value.microsecond,
            )
    return value


def format_cell(value: Any) -> Any:
    """Normalise values for Excel while keeping dates/numbers typed."""
    if _is_missing(value):
        return ""
    if hasattr(value, "tolist") and not isinstance(value, (str, bytes, datetime, date)):
        return _format_sequence(value)
    if isinstance(value, (list, tuple)):
        return _format_sequence(value)
    if hasattr(value, "item") and hasattr(value, "dtype"):
        try:
            value = value.item()
        except (TypeError, ValueError):
            pass
        if _is_missing(value):
            return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, datetime):
        return _excel_safe_datetime(value)
    if isinstance(value, date):
        return value
    if isinstance(value, Decimal):
        as_float = float(value)
        if as_float.is_integer():
            return int(as_float)
        return as_float
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return value
    return value


def _display_width_text(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        try:
            t = value.time()
        except (ValueError, OSError):
            return ""
        if t.hour or t.minute or t.second or t.microsecond:
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value)


def _excel_number_format(value: Any) -> str | None:
    if isinstance(value, datetime):
        try:
            t = value.time()
        except (ValueError, OSError):
            return "YYYY-MM-DD"
        if t.hour or t.minute or t.second or t.microsecond:
            return "YYYY-MM-DD HH:MM:SS"
        return "YYYY-MM-DD"
    if isinstance(value, date):
        return "YYYY-MM-DD"
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return "#,##0"
    if isinstance(value, float):
        return "#,##0.00"
    return None


def append_data_row(ws: Worksheet, values: Sequence[Any]) -> None:
    """
    Stream one data row immediately (write-only safe).

    Prefer raw values for speed on large multi-programme exports. Excel Table
    Style Light 1 supplies the cell grid/outlines. WriteOnlyCell is only used
    for wrapped text or date/time values.
    """
    row_idx = int(getattr(ws, "_export_data_row", 0) or 0)
    setattr(ws, "_export_data_row", row_idx + 1)
    wrap_cols: set[int] = getattr(ws, "_export_wrap_cols", set()) or set()

    cells: list[Any] = []
    for col_idx, value in enumerate(values):
        wrap = col_idx in wrap_cols
        # Skip number formats for int/float — raw values are far cheaper at scale.
        number_format = (
            _excel_number_format(value)
            if isinstance(value, (datetime, date)) and not isinstance(value, bool)
            else None
        )
        if number_format is None and not wrap:
            cells.append(value)
            continue
        cell = WriteOnlyCell(ws, value=value)
        if number_format is not None:
            cell.number_format = number_format
        if wrap:
            cell.alignment = _WRAP_ALIGN
            cell.font = _DATA_FONT
        cells.append(cell)
    ws.append(cells)


_THIN_BORDER = Border(
    left=Side(style="thin", color=_BORDER_GREY),
    right=Side(style="thin", color=_BORDER_GREY),
    top=Side(style="thin", color=_BORDER_GREY),
    bottom=Side(style="thin", color=_BORDER_GREY),
)
_HEADER_FILL = PatternFill(fill_type="solid", fgColor=_BRAND_MAROON)
_HEADER_FONT = Font(
    name=_FONT_NAME, size=_FONT_SIZE, color="FFFFFF", bold=True
)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATA_FONT = Font(name=_FONT_NAME, size=_FONT_SIZE, color="000000")
_DATA_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)
_WRAP_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
_SUSPENDED_FILL = PatternFill(fill_type="solid", fgColor=_SUSPENDED_GREY)
# Row striping + cell outlines come from Excel table style (fast at scale).
_TABLE_STYLE = TableStyleInfo(
    name="TableStyleLight1",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
# Excel table column names cannot contain these characters.
_TABLE_HEADER_BAD_CHARS = str.maketrans(
    {
        "\\": "_",
        "/": "_",
        "?": "_",
        "*": "_",
        "[": "_",
        "]": "_",
    }
)


def _excel_table_headers(headers: Sequence[str]) -> list[str]:
    """
    Return header labels that are valid unique Excel table column names.

    Must be applied both when writing the header row and when building the
    Table definition — Excel removes tables when these do not match exactly.
    """
    cleaned: list[str] = []
    seen: set[str] = set()
    for idx, raw in enumerate(headers):
        name = str(raw or "").strip() or f"Column{idx + 1}"
        name = name.translate(_TABLE_HEADER_BAD_CHARS)
        # Excel treats table column names as case-insensitive uniques.
        key = name.casefold()
        if key in seen:
            suffix = 2
            while f"{name}_{suffix}".casefold() in seen:
                suffix += 1
            name = f"{name}_{suffix}"
            key = name.casefold()
        seen.add(key)
        cleaned.append(name)
    return cleaned


def _is_wrap_header(header: str) -> bool:
    """Long-text columns: notes, comments, names (not short codes)."""
    h = str(header or "").casefold().strip()
    if not h:
        return False
    if any(token in h for token in ("note", "comment", "content", "remark")):
        return True
    if h.endswith(" name") or h.endswith("_name") or h.endswith("fullname"):
        return True
    if h in {
        "assessment",
        "module",
        "programme",
        "student",
        "course",
        "course_display_name",
        "user_full_name",
    }:
        return True
    return False


def _is_status_column(header: str) -> bool:
    """True for status fields that may contain Suspended (not Mark Status)."""
    h = str(header or "").casefold().strip()
    if h == "status":
        return True
    if h.endswith(" status") and "mark" not in h:
        return True
    return False


def _max_width_for_header(header: str) -> int:
    return MAX_WRAP_COLUMN_WIDTH if _is_wrap_header(header) else MAX_COLUMN_WIDTH


def _heuristic_column_width(header: str) -> int:
    """
    Write-only sheets must set column widths before the first append.
    Use header-based heuristics (true content autofit would require buffering).
    """
    max_w = _max_width_for_header(header)
    h = str(header or "").casefold().strip()

    # Explicit widths for common gradebook columns.
    if h == "programme":
        return 12
    if h == "email":
        return min(34, max_w)
    if h in {"module code", "course code"}:
        return min(24, max_w)
    if h in {
        "due date",
        "submitted date",
        "effective deadline",
        "last moodle access",
        "note timestamp",
    } or ("due" in h and "date" in h) or h.endswith(" deadline"):
        return min(22, max_w)

    if _is_wrap_header(header):
        return max_w
    if any(
        token in h
        for token in (
            " no",
            "number",
            "type",
            "status",
            "mark",
            "grade",
            "hours",
            "days",
            "count",
            "students",
            "modules",
        )
    ):
        return max(MIN_COLUMN_WIDTH, min(len(header) + 4, 18))
    return max(MIN_COLUMN_WIDTH, min(max(len(header) + 4, 16), max_w))


def _configure_sheet_chrome(ws: Worksheet) -> None:
    """Hide gridlines and freeze header row (must run before cells in write-only)."""
    try:
        ws.sheet_view.showGridLines = False
    except Exception:
        pass
    try:
        ws.freeze_panes = "A2"
    except Exception:
        pass


def _apply_column_widths(
    ws: Worksheet, headers: Sequence[str], col_widths: Sequence[int] | None
) -> None:
    """Set column widths before any cells are written (required for write-only)."""
    widths = list(col_widths) if col_widths else []
    for idx, header in enumerate(headers):
        max_w = _max_width_for_header(header)
        header_width = min(len(header) + 2, max_w)
        sampled = (
            widths[idx] if idx < len(widths) else _heuristic_column_width(header)
        )
        width = min(max(header_width, sampled, MIN_COLUMN_WIDTH), max_w)
        ws.column_dimensions[get_column_letter(idx + 1)].width = width


def write_headers(ws: Worksheet, headers: Sequence[str]) -> list[str]:
    """
    Write styled header row immediately and prepare streaming export state.

    Column widths are applied before the first append (write-only requirement).
    """
    _configure_sheet_chrome(ws)
    safe_headers = _excel_table_headers(headers)
    setattr(ws, "_export_headers", safe_headers)
    setattr(ws, "_export_data_row", 0)
    setattr(
        ws,
        "_export_wrap_cols",
        {i for i, h in enumerate(safe_headers) if _is_wrap_header(h)},
    )
    setattr(
        ws,
        "_export_status_cols",
        {i for i, h in enumerate(safe_headers) if _is_status_column(h)},
    )

    heuristic_widths = [_heuristic_column_width(h) for h in safe_headers]
    _apply_column_widths(ws, safe_headers, heuristic_widths)

    cells: list[Any] = []
    for header in safe_headers:
        cell = WriteOnlyCell(ws, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
        cells.append(cell)
    ws.append(cells)
    try:
        ws.row_dimensions[1].height = 22
    except Exception:
        pass
    return safe_headers


def _unique_table_name(ws: Worksheet) -> str:
    """Build a workbook-unique Excel table display name from the sheet title."""
    base = "".join(ch for ch in (ws.title or "Sheet") if ch.isalnum()) or "Sheet"
    if not base[0].isalpha():
        base = f"T{base}"
    base = base[:40]
    existing: set[str] = set()
    wb = ws.parent
    if wb is not None:
        for sheet in wb.worksheets:
            tables = getattr(sheet, "tables", None)
            if tables is None:
                continue
            try:
                existing.update(tables.keys())
            except Exception:
                continue
    if base not in existing:
        return base
    idx = 2
    while f"{base}_{idx}" in existing:
        idx += 1
    return f"{base}_{idx}"


def _apply_suspended_conditional_format(
    ws: Worksheet, headers: Sequence[str], data_row_count: int
) -> None:
    """Grey-fill cells whose status value contains 'suspended'."""
    if data_row_count < 1:
        return
    last_row = data_row_count + 1
    for idx, header in enumerate(headers):
        if not _is_status_column(header):
            continue
        col = get_column_letter(idx + 1)
        range_ref = f"{col}2:{col}{last_row}"
        ws.conditional_formatting.add(
            range_ref,
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("suspended",{col}2))'],
                fill=_SUSPENDED_FILL,
                font=_DATA_FONT,
            ),
        )


def finish_sheet(
    ws: Worksheet,
    headers: Sequence[str],
    data_row_count: int,
    col_widths: Sequence[int] | None = None,
) -> None:
    """
    Finalize a streamed sheet: Excel Table (when data exists), filters, and
    Suspended conditional formatting.

    Headers and data rows are already written by write_headers / append_data_row.
    ``col_widths`` is accepted for caller compatibility but widths were applied
    heuristically up-front (write-only constraint).
    """
    del col_widths  # widths already applied in write_headers for write-only
    _configure_sheet_chrome(ws)
    safe_headers = list(
        getattr(ws, "_export_headers", None) or _excel_table_headers(headers)
    )
    column_count = len(safe_headers)
    if column_count < 1:
        return

    streamed = int(getattr(ws, "_export_data_row", 0) or 0)
    if streamed:
        data_row_count = streamed

    last_col = get_column_letter(column_count)
    last_row = max(1, data_row_count + 1)
    ref = f"A1:{last_col}{last_row}"

    if data_row_count >= 1:
        # Always format as an Excel Table so row striping (and filter UI) apply,
        # including on large Student Assessment Detail sheets.
        table = Table(displayName=_unique_table_name(ws), ref=ref)
        table.totalsRowShown = False
        table.tableStyleInfo = _TABLE_STYLE
        table._initialise_columns()
        for column, header in zip(table.tableColumns, safe_headers):
            column.name = header
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="In write-only mode you must add table columns manually",
                category=UserWarning,
            )
            ws.add_table(table)
    else:
        ws.auto_filter.ref = ref

    _apply_suspended_conditional_format(ws, safe_headers, data_row_count)


def _update_col_widths(
    widths: list[int],
    values: Sequence[Any],
    sample_remaining: list[int] | None = None,
    headers: Sequence[str] | None = None,
) -> None:
    """Track max display width per column (optional; ignored by write-only finish)."""
    if sample_remaining is not None:
        if sample_remaining[0] <= 0:
            return
        sample_remaining[0] -= 1
    for idx, value in enumerate(values):
        text = _display_width_text(value)
        max_w = (
            _max_width_for_header(headers[idx])
            if headers is not None and idx < len(headers)
            else MAX_COLUMN_WIDTH
        )
        length = min(len(text) + 2, max_w)
        if idx >= len(widths):
            widths.append(max(MIN_COLUMN_WIDTH, length))
        else:
            widths[idx] = max(widths[idx], length)


def add_header_only_sheet(wb: Workbook, title: str, headers: Sequence[str]) -> None:
    """Create a mart sheet with column headers only (no data rows)."""
    ws = wb.create_sheet(title=title[:31])
    write_headers(ws, headers)
    finish_sheet(ws, headers, 0)


def write_mapped_rows(
    ws: Worksheet,
    rows: Iterator[dict[str, Any]] | Sequence[dict[str, Any]],
    headers: Sequence[str],
    field_map: Sequence[tuple[str, tuple[str, ...]]],
) -> int:
    write_headers(ws, headers)
    count = 0
    for raw in rows:
        row = normalize_row(raw)
        values = [format_cell(pick(row, *aliases)) for _, aliases in field_map]
        append_data_row(ws, values)
        count += 1
    finish_sheet(ws, headers, count)
    return count


def _build_mart_filter_sql(
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    table: str,
    programme_codes: Sequence[str],
    category_name: str | None,
    order_columns: Sequence[str] | None = None,
    *,
    limit: int | None = None,
) -> tuple[str, list[Any]] | None:
    """
    Build SELECT for a mart filter. Returns None when the table/columns are missing.
    """
    codes = [
        str(c).strip().upper()
        for c in programme_codes
        if str(c).strip()
    ]
    if not codes:
        return None

    spec = MART_FILTER.get(table, {"programme_cols": ("programme",), "category_cols": ()})
    programme_candidates: tuple[str, ...] = tuple(
        spec.get("programme_cols", ("programme",))
    )
    category_candidates: tuple[str, ...] = tuple(spec.get("category_cols", ()))
    programme_ilike: bool = bool(spec.get("programme_ilike"))
    programme_array: bool = bool(spec.get("programme_array"))

    relation = qualified_relation(schema, table)
    try:
        mart_cols = _mart_columns(conn, schema, table)
    except duckdb.CatalogException:
        return None
    programme_col = _pick_first_mart_column(mart_cols, programme_candidates)
    if not programme_col:
        return None

    where_parts: list[str] = []
    params: list[Any] = []

    for programme_code in codes:
        if programme_array:
            where_parts.append(
                f'EXISTS (SELECT 1 FROM unnest("{programme_col}") AS _md_prog(prog) '
                f'WHERE UPPER(TRIM(CAST(_md_prog.prog AS VARCHAR))) = UPPER(TRIM(?)))'
            )
            params.append(programme_code)
        elif programme_ilike:
            where_parts.append(f'CAST("{programme_col}" AS VARCHAR) ILIKE ?')
            params.append(f"%{programme_code}%")
        else:
            where_parts.append(
                f'UPPER(TRIM(CAST("{programme_col}" AS VARCHAR))) = UPPER(TRIM(?))'
            )
            params.append(programme_code)

    programme_clause = (
        where_parts[0] if len(where_parts) == 1 else "(" + " OR ".join(where_parts) + ")"
    )
    query = f"SELECT * FROM {relation} WHERE {programme_clause}"
    if category_name and category_candidates:
        category_col = _pick_first_mart_column(mart_cols, category_candidates)
        if category_col:
            query += f' AND TRIM(CAST("{category_col}" AS VARCHAR)) = TRIM(?)'
            params.append(category_name)
    if order_columns:
        present = [col for col in order_columns if col.lower() in mart_cols]
        if present:
            order = ", ".join(f'"{mart_cols[col.lower()]}"' for col in present)
            query += f" ORDER BY {order}"
        else:
            query += " ORDER BY 1"
    else:
        query += " ORDER BY 1"
    if limit is not None:
        query += f" LIMIT {int(limit)}"
    return query, params


def iter_mart_rows(
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    table: str,
    programme_codes: Sequence[str],
    category_name: str | None,
    order_columns: Sequence[str] | None = None,
    *,
    chunk_size: int = FETCH_CHUNK_SIZE,
) -> Iterator[dict[str, Any]]:
    """Stream mart rows in chunks (no pandas DataFrame)."""
    built = _build_mart_filter_sql(
        conn,
        schema,
        table,
        programme_codes,
        category_name,
        order_columns,
    )
    if not built:
        return
    query, params = built
    result = conn.execute(query, params)
    columns = [str(desc[0]) for desc in result.description]
    while True:
        batch = result.fetchmany(chunk_size)
        if not batch:
            break
        for tup in batch:
            yield normalize_row(dict(zip(columns, tup)))


def fetch_mart_rows(
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    table: str,
    programme_codes: Sequence[str],
    category_name: str | None,
    order_columns: Sequence[str] | None = None,
) -> list[dict[str, Any]]:
    """Materialise mart rows (prefer iter_mart_rows for large sheets)."""
    return list(
        iter_mart_rows(
            conn,
            schema,
            table,
            programme_codes,
            category_name,
            order_columns,
        )
    )


def mart_has_rows(
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    table: str,
    programme_codes: Sequence[str],
    category_name: str | None,
) -> bool:
    """True when at least one matching mart row exists (LIMIT 1)."""
    built = _build_mart_filter_sql(
        conn,
        schema,
        table,
        programme_codes,
        category_name,
        order_columns=None,
        limit=1,
    )
    if not built:
        return False
    query, params = built
    return conn.execute(query, params).fetchone() is not None


def count_suspended_students(
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    programme_codes: Sequence[str],
    category_name: str | None,
) -> int:
    built = _build_mart_filter_sql(
        conn,
        schema,
        TABLE_STUDENT,
        programme_codes,
        category_name,
        order_columns=None,
    )
    if not built:
        return 0
    base_query, params = built
    try:
        student_cols = _mart_columns(conn, schema, TABLE_STUDENT)
    except duckdb.CatalogException:
        return 0
    status_col = student_cols.get("status")
    if not status_col:
        return 0
    # Drop ORDER BY for the aggregate wrapper (cheaper + avoids nested ORDER issues).
    base_no_order = base_query.rsplit(" ORDER BY ", 1)[0]
    query = f"""
        SELECT COUNT(*) FROM ({base_no_order}) AS _students
        WHERE LOWER(TRIM(CAST("{status_col}" AS VARCHAR))) = 'suspended'
    """
    try:
        row = conn.execute(query, params).fetchone()
    except Exception:
        return 0
    return int(row[0] or 0) if row else 0


def count_submitted_late_by_keys(
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    programme_codes: Sequence[str],
    category_name: str | None,
    group_key_candidates: Sequence[Sequence[str]],
) -> dict[tuple[str, ...], int]:
    """
    Count assessment-detail rows where status = 'submitted_late', grouped by keys.

    group_key_candidates: for each group dimension, ordered column-name candidates
    (first present column wins), e.g. (("programme", "course_prefix"), ("student_no",)).
    """
    built = _build_mart_filter_sql(
        conn,
        schema,
        TABLE_ASSESSMENT,
        programme_codes,
        category_name,
        order_columns=None,
    )
    if not built:
        return {}
    base_query, params = built
    try:
        mart_cols = _mart_columns(conn, schema, TABLE_ASSESSMENT)
    except duckdb.CatalogException:
        return {}

    status_col = _pick_first_mart_column(mart_cols, ("status",))
    if not status_col:
        return {}

    group_cols: list[str] = []
    for candidates in group_key_candidates:
        col = _pick_first_mart_column(mart_cols, candidates)
        if not col:
            return {}
        group_cols.append(col)

    base_no_order = base_query.rsplit(" ORDER BY ", 1)[0]
    select_keys = ", ".join(f'"{c}"' for c in group_cols)
    group_sql = ", ".join(f'"{c}"' for c in group_cols)
    query = f"""
        SELECT {select_keys}, COUNT(*) AS late_count
        FROM ({base_no_order}) AS _detail
        WHERE LOWER(REPLACE(REPLACE(TRIM(CAST("{status_col}" AS VARCHAR)),
              ' ', '_'), '-', '_')) = 'submitted_late'
        GROUP BY {group_sql}
    """
    try:
        rows = conn.execute(query, params).fetchall()
    except Exception:
        return {}

    out: dict[tuple[str, ...], int] = {}
    for tup in rows:
        *keys, count = tup
        key = tuple(str(k or "").strip().upper() for k in keys)
        out[key] = int(count or 0)
    return out


def _course_notes_table_columns(
    conn: duckdb.DuckDBPyConnection, schema: str
) -> list[str]:
    try:
        relation = qualified_relation(schema, TABLE_COURSE_NOTES)
        described = conn.execute(
            f"DESCRIBE SELECT * FROM {relation} LIMIT 0"
        ).fetchall()
        columns = [str(row[0]) for row in described]
        if columns:
            return columns
    except Exception:
        pass
    return list(COURSE_NOTES_COLUMNS)


def iter_course_note_rows(
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    programme_codes: Sequence[str],
    category_name: str | None,
    *,
    chunk_size: int = FETCH_CHUNK_SIZE,
) -> Iterator[dict[str, Any]]:
    """Stream course notes for students in the offering (SQL subquery, no student preload)."""
    student_filter = _build_mart_filter_sql(
        conn,
        schema,
        TABLE_STUDENT,
        programme_codes,
        category_name,
        order_columns=None,
    )
    if not student_filter:
        return
    student_sql, student_params = student_filter
    try:
        relation = qualified_relation(schema, TABLE_COURSE_NOTES)
        note_cols = {
            str(row[0]).lower(): str(row[0])
            for row in conn.execute(
                f"DESCRIBE SELECT * FROM {relation} LIMIT 0"
            ).fetchall()
        }
        username_col = note_cols.get("username")
        if not username_col:
            return
        student_cols = _mart_columns(conn, schema, TABLE_STUDENT)
        student_no_col = student_cols.get("student_no")
        if not student_no_col:
            return
        programme_col = (
            student_cols.get("programme")
            or student_cols.get("program_code")
            or student_cols.get("course_prefix")
        )
        programme_select = (
            f'CAST(s."{programme_col}" AS VARCHAR) AS programme'
            if programme_col
            else "CAST(NULL AS VARCHAR) AS programme"
        )
        student_base_sql = student_sql.rsplit(" ORDER BY ", 1)[0]
        note_select_cols = ", ".join(
            f'n."{col}"'
            for key, col in note_cols.items()
            if key != "programme"
        )
        if "timestamp" in note_cols:
            order_sql = (
                f's.programme, n."{note_cols["timestamp"]}" DESC, '
                f'n."{username_col}", '
                f'n."{note_cols.get("course_display_name", username_col)}"'
            )
        else:
            order_sql = f's.programme, n."{username_col}"'
        select_list = (
            f"s.programme AS programme, {note_select_cols}"
            if note_select_cols
            else "s.programme AS programme"
        )
        query = f"""
            SELECT {select_list}
            FROM {relation} AS n
            INNER JOIN (
                SELECT
                    TRIM(CAST(s."{student_no_col}" AS VARCHAR)) AS student_no,
                    {programme_select}
                FROM ({student_base_sql}) AS s
                WHERE TRIM(CAST(s."{student_no_col}" AS VARCHAR)) <> ''
            ) AS s
                ON TRIM(CAST(n."{username_col}" AS VARCHAR)) = s.student_no
            ORDER BY {order_sql}
        """
        result = conn.execute(query, student_params)
        columns = [str(desc[0]) for desc in result.description]
        while True:
            batch = result.fetchmany(chunk_size)
            if not batch:
                break
            for tup in batch:
                yield dict(zip(columns, tup))
    except Exception:
        return


def fetch_course_note_rows(
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    programme_codes: Sequence[str],
    category_name: str | None,
) -> list[dict[str, Any]]:
    return list(
        iter_course_note_rows(conn, schema, programme_codes, category_name)
    )


def write_programme_summary(
    ws: Worksheet,
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    programme_codes: Sequence[str],
    category_name: str | None,
    display_programme_code: str = "",
) -> None:
    """Write Programme Summary as a table: one row per programme."""
    headers = list(PROGRAMME_SUMMARY_HEADERS)
    write_headers(ws, headers)
    widths = [max(12, min(len(h) + 2, MAX_COLUMN_WIDTH)) for h in headers]
    sample_remaining = [COLUMN_WIDTH_SAMPLE_ROWS]

    metric_aliases: list[tuple[str, tuple[str, ...]]] = [
        ("Programme", ("programme", "program_code")),
        ("Students", ("students",)),
        ("Suspended Students", ("suspended_students", "suspended_student_count")),
        ("Active Modules", ("active_modules",)),
        ("Submitted Assessments", ("submitted_assessments",)),
        ("Missed Assessments", ("missed_assessments",)),
        ("Late Submissions", ("late_submissions",)),
        ("Upcoming Deadlines (14 Days)", ("upcoming_deadlines_14_days",)),
        ("Students With Missed Assessments", ("students_with_missed_assessments",)),
    ]

    rows = list(
        iter_mart_rows(
            conn,
            schema,
            TABLE_PROGRAMME,
            programme_codes,
            category_name,
            order_columns=["programme"],
        )
    )
    late_by_programme = count_submitted_late_by_keys(
        conn,
        schema,
        programme_codes,
        category_name,
        (("programme", "course_prefix", "program_code"),),
    )

    # Ensure every selected programme appears, even if mart summary is missing.
    seen_codes: set[str] = set()
    count = 0
    for raw in rows:
        row = normalize_row(raw)
        programme_value = pick(row, "programme", "program_code")
        programme_code = str(programme_value or "").strip().upper()
        if programme_code:
            seen_codes.add(programme_code)
        values: list[Any] = []
        for label, aliases in metric_aliases:
            value = pick(row, *aliases)
            if label == "Suspended Students" and value == "":
                value = count_suspended_students(
                    conn,
                    schema,
                    [programme_code] if programme_code else programme_codes,
                    category_name,
                )
            if label == "Late Submissions":
                value = late_by_programme.get(
                    (programme_code,), value if value != "" else 0
                )
            if label == "Programme" and not value:
                value = (
                    display_programme_code
                    or programme_code
                    or (programme_codes[0] if programme_codes else "")
                )
            values.append(format_cell(value))
        _update_col_widths(widths, values, sample_remaining)
        append_data_row(ws, values)
        count += 1

    for code in programme_codes:
        code_norm = str(code).strip().upper()
        if not code_norm or code_norm in seen_codes:
            continue
        suspended = count_suspended_students(
            conn, schema, [code_norm], category_name
        )
        late = late_by_programme.get((code_norm,), 0)
        values = [
            format_cell(code_norm),
            "",
            format_cell(suspended),
            "",
            "",
            "",
            format_cell(late),
            "",
            "",
        ]
        _update_col_widths(widths, values, sample_remaining)
        append_data_row(ws, values)
        count += 1

    if count == 0 and (display_programme_code or programme_codes):
        code = (
            display_programme_code.strip().upper()
            if display_programme_code
            else str(programme_codes[0]).strip().upper()
        )
        values = [format_cell(code), "", "", "", "", "", "", "", ""]
        _update_col_widths(widths, values, sample_remaining)
        append_data_row(ws, values)
        count = 1

    finish_sheet(ws, headers, count, widths)


def write_student_summary(
    ws: Worksheet,
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    programme_codes: Sequence[str],
    category_name: str | None,
) -> None:
    headers = [
        "Programme",
        "Student No",
        "Student",
        "Email",
        "Status",
        "Modules",
        "Missed Submissions",
        "Late Submissions",
        "Upcoming Submissions",
        "Last Moodle Access",
        "Days Since Access",
        *[label for label, _ in NOTE_FIELD_MAP],
    ]
    field_map: list[tuple[str, tuple[str, ...]]] = [
        ("Programme", ("programme", "program_code")),
        ("Student No", ("student_no",)),
        ("Student", ("student",)),
        ("Email", ("email",)),
        ("Status", ("status",)),
        ("Modules", ("total_modules", "modules")),
        ("Missed Submissions", ("missed_submissions", "missed")),
        ("Late Submissions", ("late_submissions", "late")),
        ("Upcoming Submissions", ("upcoming_submissions", "upcoming")),
        ("Last Moodle Access", ("last_moodle_access",)),
        ("Days Since Access", ("days_since_access",)),
        *NOTE_FIELD_MAP,
    ]
    late_by_student = count_submitted_late_by_keys(
        conn,
        schema,
        programme_codes,
        category_name,
        (
            ("programme", "course_prefix", "program_code"),
            ("student_no",),
        ),
    )
    write_headers(ws, headers)
    widths = [max(12, min(len(h) + 2, MAX_COLUMN_WIDTH)) for h in headers]
    sample_remaining = [COLUMN_WIDTH_SAMPLE_ROWS]
    count = 0
    for raw in iter_mart_rows(
        conn,
        schema,
        TABLE_STUDENT,
        programme_codes,
        category_name,
        order_columns=["programme", "student_no"],
    ):
        row = normalize_row(raw)
        values: list[Any] = []
        for label, aliases in field_map:
            value = pick(row, *aliases)
            if label == "Late Submissions":
                prog = str(pick(row, "programme", "program_code") or "").strip().upper()
                student_no = str(pick(row, "student_no") or "").strip().upper()
                value = late_by_student.get((prog, student_no), value if value != "" else 0)
            values.append(format_cell(value))
        _update_col_widths(widths, values, sample_remaining)
        append_data_row(ws, values)
        count += 1
    finish_sheet(ws, headers, count, widths)


def write_module_summary(
    ws: Worksheet,
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    programme_codes: Sequence[str],
    category_name: str | None,
) -> None:
    headers = [
        "Programme",
        "Module Code",
        "Module",
        "Students",
        "Submitted",
        "Missed Submissions",
        "Late Submissions",
        "Upcoming",
    ]
    field_map = [
        ("Programme", ("programme", "program_code")),
        ("Module Code", ("module_code",)),
        ("Module", ("module",)),
        ("Students", ("students",)),
        ("Submitted", ("total_submissions", "submitted")),
        ("Missed Submissions", ("missed_submissions", "missed")),
        ("Late Submissions", ("late_submissions", "late")),
        ("Upcoming", ("upcoming_assessments", "upcoming")),
    ]
    late_by_module = count_submitted_late_by_keys(
        conn,
        schema,
        programme_codes,
        category_name,
        (
            ("programme", "course_prefix", "program_code"),
            ("course_shortname", "module_code"),
        ),
    )
    write_headers(ws, headers)
    widths = [max(12, min(len(h) + 2, MAX_COLUMN_WIDTH)) for h in headers]
    sample_remaining = [COLUMN_WIDTH_SAMPLE_ROWS]
    count = 0
    for raw in iter_mart_rows(
        conn,
        schema,
        TABLE_MODULE,
        programme_codes,
        category_name,
        order_columns=["programme", "module_code", "module"],
    ):
        row = normalize_row(raw)
        values: list[Any] = []
        for label, aliases in field_map:
            value = pick(row, *aliases)
            if label == "Late Submissions":
                prog = str(pick(row, "programme", "program_code") or "").strip().upper()
                module_code = str(
                    pick(row, "module_code", "course_shortname") or ""
                ).strip().upper()
                value = late_by_module.get(
                    (prog, module_code), value if value != "" else 0
                )
            values.append(format_cell(value))
        _update_col_widths(widths, values, sample_remaining)
        append_data_row(ws, values)
        count += 1
    finish_sheet(ws, headers, count, widths)


def write_submission_trends(
    ws: Worksheet,
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    programme_codes: Sequence[str],
    category_name: str | None,
) -> None:
    headers = [
        "Programme",
        "Module Code",
        "Module",
        "Assessment",
        "Assessment Type",
        "Due Date",
        "Total Students",
        "Submitted Count",
        "Missed Count",
        "Late Submissions",
    ]
    field_map = [
        ("Programme", ("programme", "program_code")),
        ("Module Code", ("module_code",)),
        ("Module", ("module",)),
        ("Assessment", ("assessment",)),
        ("Assessment Type", ("assessment_type",)),
        ("Due Date", ("effective_deadline_at", "due_date")),
        ("Total Students", ("total_students",)),
        ("Submitted Count", ("submitted_count",)),
        ("Missed Count", ("missed_count",)),
        ("Late Submissions", ("late_count", "late_submissions")),
    ]
    late_by_assessment = count_submitted_late_by_keys(
        conn,
        schema,
        programme_codes,
        category_name,
        (
            ("programme", "course_prefix", "program_code"),
            ("course_shortname", "module_code"),
            ("assessment_name", "assessment"),
        ),
    )
    write_headers(ws, headers)
    widths = [max(12, min(len(h) + 2, MAX_COLUMN_WIDTH)) for h in headers]
    sample_remaining = [COLUMN_WIDTH_SAMPLE_ROWS]
    count = 0
    for raw in iter_mart_rows(
        conn,
        schema,
        TABLE_TRENDS,
        programme_codes,
        category_name,
        order_columns=["programme", "module_code", "assessment"],
    ):
        row = normalize_row(raw)
        values: list[Any] = []
        for label, aliases in field_map:
            value = pick(row, *aliases)
            if label == "Late Submissions":
                prog = str(pick(row, "programme", "program_code") or "").strip().upper()
                module_code = str(
                    pick(row, "module_code", "course_shortname") or ""
                ).strip().upper()
                assessment = str(
                    pick(row, "assessment", "assessment_name") or ""
                ).strip().upper()
                value = late_by_assessment.get(
                    (prog, module_code, assessment), value if value != "" else 0
                )
            values.append(format_cell(value))
        _update_col_widths(widths, values, sample_remaining)
        append_data_row(ws, values)
        count += 1
    finish_sheet(ws, headers, count, widths)


def write_student_assessment_detail(
    ws: Worksheet,
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    programme_codes: Sequence[str],
    category_name: str | None,
) -> None:
    headers = [
        "Programme",
        "Student No",
        "Student",
        "Email",
        "Module Code",
        "Module",
        "Assessment",
        "Assessment Type",
        "Due Date",
        "Submitted Date",
        "Status",
        "Mark Status",
        "Mark",
        "Max Grade",
        *[label for label, _ in NOTE_FIELD_MAP],
    ]
    write_headers(ws, headers)
    widths = [max(12, min(len(h) + 2, MAX_COLUMN_WIDTH)) for h in headers]
    sample_remaining = [COLUMN_WIDTH_SAMPLE_ROWS]
    count = 0
    for raw in iter_mart_rows(
        conn,
        schema,
        TABLE_ASSESSMENT,
        programme_codes,
        category_name,
        order_columns=[
            "programme",
            "course_prefix",
            "course_shortname",
            "assessment_name",
            "student_no",
        ],
    ):
        row = normalize_row(raw)
        submitted = pick(
            row,
            "grade_submitted_at",
            "last_attempt_at",
            "submitted_at",
            "graded_at",
        )
        # Mart often keeps status='missed' for manual/offline grades when
        # is_submitted/has_attempt are false — treat a present grade as graded.
        raw_status = (
            str(pick(row, "status") or "")
            .strip()
            .lower()
            .replace(" ", "_")
            .replace("-", "_")
        )
        grade = pick(row, "grade_raw")
        graded_at = pick(row, "graded_at")
        has_grade = (grade != "" and grade is not None) or (
            graded_at != "" and graded_at is not None
        )
        if raw_status in {"submitted", "graded", "submitted_late"}:
            display_status = raw_status
        elif has_grade:
            display_status = "graded"
        else:
            display_status = pick(row, "status")
        is_submitted = row.get("is_submitted")
        if (
            not has_grade
            and is_submitted not in (True, 1)
            and str(is_submitted).lower()
            not in {
                "true",
                "t",
                "1",
                "yes",
            }
        ):
            submitted = ""
        values = [
            format_cell(pick(row, "programme", "course_prefix", "program_code")),
            format_cell(pick(row, "student_no", "user_username")),
            format_cell(pick(row, "user_fullname")),
            format_cell(pick(row, "user_email")),
            format_cell(pick(row, "course_shortname")),
            format_cell(pick(row, "course_fullname")),
            format_cell(pick(row, "assessment", "assessment_name")),
            format_cell(pick(row, "assessment_type")),
            format_cell(pick(row, "due_at", "effective_deadline_at")),
            format_cell(submitted),
            format_cell(display_status),
            format_cell(pick(row, "mark_status")),
            format_cell(pick(row, "grade_raw")),
            format_cell(pick(row, "max_grade")),
            *[format_cell(pick(row, *aliases)) for _, aliases in NOTE_FIELD_MAP],
        ]
        _update_col_widths(widths, values, sample_remaining)
        append_data_row(ws, values)
        count += 1
    finish_sheet(ws, headers, count, widths)


def write_missed_assessments(
    ws: Worksheet,
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    programme_codes: Sequence[str],
    category_name: str | None,
) -> None:
    headers = [
        "Programme",
        "Student No",
        "Student",
        "Email",
        "Module Code",
        "Module",
        "Assessment",
        "Assessment Type",
        "Due Date",
        "Effective Deadline",
        "Days Overdue",
        "Status",
        "Mark Status",
        *[label for label, _ in NOTE_FIELD_MAP],
    ]
    field_map = [
        ("Programme", ("programme", "program_code")),
        ("Student No", ("student_no",)),
        ("Student", ("student",)),
        ("Email", ("email",)),
        ("Module Code", ("module_code",)),
        ("Module", ("module",)),
        ("Assessment", ("assessment",)),
        ("Assessment Type", ("assessment_type",)),
        ("Due Date", ("due_date", "effective_deadline_at")),
        ("Effective Deadline", ("effective_deadline_at", "due_date")),
        ("Days Overdue", ("days_overdue",)),
        ("Status", ("status",)),
        ("Mark Status", ("mark_status",)),
        *NOTE_FIELD_MAP,
    ]
    write_mapped_rows(
        ws,
        iter_mart_rows(
            conn,
            schema,
            TABLE_MISSED,
            programme_codes,
            category_name,
            order_columns=["programme", "days_overdue", "student_no"],
        ),
        headers,
        field_map,
    )


def write_gradebook_course_notes(
    ws: Worksheet,
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    programme_codes: Sequence[str],
    category_name: str | None,
) -> None:
    note_columns = _course_notes_table_columns(conn, schema)
    # Programme first for filtering; avoid duplicating if the mart already has it.
    note_columns_no_prog = [
        col for col in note_columns if str(col).lower() != "programme"
    ]
    headers = ["Programme", *note_columns_no_prog]
    write_headers(ws, headers)
    widths = [max(12, min(len(h) + 2, MAX_COLUMN_WIDTH)) for h in headers]
    sample_remaining = [COLUMN_WIDTH_SAMPLE_ROWS]
    count = 0
    for raw in iter_course_note_rows(
        conn, schema, programme_codes, category_name
    ):
        row_norm = normalize_row(raw)
        values = [format_cell(pick(row_norm, "programme"))]
        for col in note_columns_no_prog:
            values.append(format_cell(row_norm.get(str(col).lower(), raw.get(col))))
        _update_col_widths(widths, values, sample_remaining)
        append_data_row(ws, values)
        count += 1
    finish_sheet(ws, headers, count, widths)


def write_upcoming_deadlines(
    ws: Worksheet,
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    programme_codes: Sequence[str],
    category_name: str | None,
) -> None:
    headers = [
        "Programme",
        "Module Code",
        "Module",
        "Assessment",
        "Assessment Type",
        "Effective Deadline",
        "Hours Until Due",
    ]
    field_map = [
        ("Programme", ("programme", "program_code")),
        ("Module Code", ("module_code",)),
        ("Module", ("module",)),
        ("Assessment", ("assessment",)),
        ("Assessment Type", ("assessment_type",)),
        ("Effective Deadline", ("effective_deadline_at",)),
        ("Hours Until Due", ("hours_until_due",)),
    ]
    write_mapped_rows(
        ws,
        iter_mart_rows(
            conn,
            schema,
            TABLE_UPCOMING,
            programme_codes,
            category_name,
            order_columns=["programme", "hours_until_due", "effective_deadline_at"],
        ),
        headers,
        field_map,
    )


def _export_codes_for_check(
    programme_codes: Sequence[str],
    display_programme_code: str,
) -> list[str]:
    """Codes to probe marts — include dropdown prefix when mapping differs from mart grain."""
    seen: set[str] = set()
    codes: list[str] = []
    for raw in (*programme_codes, display_programme_code):
        code = str(raw).strip().upper()
        if code and code not in seen:
            seen.add(code)
            codes.append(code)
    return codes


def _export_fallback_enabled() -> bool:
    flag = (read_env_value("WAREHOUSE_EXPORT_FALLBACK") or "true").lower()
    return flag not in ("0", "false", "no", "off")


def offering_has_gradebook_rows(
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    programme_codes: Sequence[str],
    category_name: str | None,
    display_programme_code: str = "",
) -> bool:
    """
    True when any gradebook mart has rows for this offering.

    Uses LIMIT 1 probes so large programmes (e.g. BCMAC) do not load full marts
    just to decide between warehouse and fallback export.
    """
    codes = _export_codes_for_check(programme_codes, display_programme_code)
    if not codes:
        return False

    probes: tuple[str, ...] = (
        TABLE_STUDENT,
        TABLE_MODULE,
        TABLE_PROGRAMME,
        TABLE_TRENDS,
        TABLE_ASSESSMENT,
        TABLE_MISSED,
        TABLE_UPCOMING,
    )
    for table in probes:
        if mart_has_rows(conn, schema, table, codes, category_name):
            return True
    # Course notes: cheap existence via subquery LIMIT 1.
    for _ in iter_course_note_rows(conn, schema, codes, category_name, chunk_size=1):
        return True
    return False


def build_workbook(
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    programme_codes: Sequence[str],
    category_name: str | None,
    output_dir: Path,
    display_programme_code: str = "",
) -> Path:
    codes = _export_codes_for_check(programme_codes, display_programme_code)
    single_display = (
        display_programme_code.strip().upper()
        if display_programme_code
        else (codes[0] if len(codes) == 1 else "")
    )

    # Fallback remains single-programme only (multi-programme uses mart combine path).
    if (
        _export_fallback_enabled()
        and category_name
        and single_display
        and len(codes) <= 1
        and not offering_has_gradebook_rows(
            conn,
            schema,
            codes,
            category_name,
            display_programme_code=single_display,
        )
    ):
        return build_workbook_offering_fallback(
            conn,
            category_name,
            single_display,
            output_dir,
            single_display,
            schema=schema,
            programme_codes=codes,
        )

    export_codes = codes or list(programme_codes)

    # write_only streams rows to disk and avoids Cell objects for every value.
    wb = Workbook(write_only=True)

    sheets: list[tuple[str, Any]] = [
        ("Programme Summary", write_programme_summary),
        ("Student Summary", write_student_summary),
        ("Module Summary", write_module_summary),
        ("Submission Trends", write_submission_trends),
        ("Student Assessment Detail", write_student_assessment_detail),
        ("Missed Assessments", write_missed_assessments),
        ("Upcoming Deadlines", write_upcoming_deadlines),
        (COURSE_NOTES_SHEET_TITLE, write_gradebook_course_notes),
    ]

    for title, writer in sheets:
        ws = wb.create_sheet(title=title[:31])
        if writer is write_programme_summary:
            writer(
                ws,
                conn,
                schema,
                export_codes,
                category_name,
                display_programme_code=single_display,
            )
        else:
            writer(ws, conn, schema, export_codes, category_name)
        # Release sheet-local buffers between large marts.
        gc.collect()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if len(export_codes) == 1:
        file_code = export_codes[0]
    else:
        joined = "_".join(export_codes)
        file_code = joined if len(joined) <= 48 else f"batch_{len(export_codes)}prog"
    safe_code = file_code.replace(" ", "_")
    out_path = output_dir / f"gradebook_{safe_code}_{timestamp}.xlsx"
    wb.save(out_path)
    return out_path.resolve()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export gradebook Excel from warehouse moodle_processed marts"
    )
    parser.add_argument(
        "--programme-code",
        action="append",
        dest="programme_codes",
        required=True,
        help="Programme code to include (repeat for multi-programme batch export)",
    )
    parser.add_argument(
        "--category-name",
        default="",
        help="Moodle category label (intake); required to scope export to one category",
    )
    parser.add_argument(
        "--warehouse-schema",
        default=None,
        help="Schema for gradebook marts (default: moodle_processed)",
    )
    parser.add_argument(
        "--output-dir",
        default=".",
        help="Directory for the generated workbook",
    )
    args = parser.parse_args()

    seen: set[str] = set()
    programme_codes: list[str] = []
    for raw in args.programme_codes or []:
        code = str(raw).strip().upper()
        if code and code not in seen:
            seen.add(code)
            programme_codes.append(code)
    if not programme_codes:
        raise SystemExit("At least one --programme-code is required")

    category_name = str(args.category_name or "").strip() or None
    schema = args.warehouse_schema or gradebook_schema() or DEFAULT_SCHEMA
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    conn = connect_motherduck()
    try:
        out_path = build_workbook(
            conn,
            schema,
            programme_codes,
            category_name,
            output_dir,
            display_programme_code=programme_codes[0] if len(programme_codes) == 1 else "",
        )
    except Exception:
        import traceback

        traceback.print_exc()
        raise
    finally:
        conn.close()

    print(out_path)


if __name__ == "__main__":
    main()
