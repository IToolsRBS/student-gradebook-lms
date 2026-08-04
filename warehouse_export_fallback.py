"""
Export fallback when gradebook marts have no rows (inactive / non-visible-recent courses).

Uses dim_courses + stg_moodle_enrollments + stg_moodle_grades for the selected
category and course shortname prefix (e.g. 2025 January + PDEML).
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Iterator, Sequence

import duckdb
from openpyxl import Workbook

from motherduck_client import (
    courses_schema,
    qualified_relation,
    staging_schema,
)
from warehouse_metadata import course_ids_for_offering

FETCH_CHUNK_SIZE = 1000


def _course_id_filter(course_ids: Sequence[int]) -> tuple[str, list[Any]]:
    placeholders = ", ".join("?" for _ in course_ids)
    return f"CAST(dc.course_id AS BIGINT) IN ({placeholders})", list(course_ids)


def _iter_query_dicts(
    conn: duckdb.DuckDBPyConnection,
    query: str,
    params: Sequence[Any],
    *,
    chunk_size: int = FETCH_CHUNK_SIZE,
) -> Iterator[dict[str, Any]]:
    result = conn.execute(query, list(params))
    columns = [str(desc[0]) for desc in result.description]
    while True:
        batch = result.fetchmany(chunk_size)
        if not batch:
            break
        for tup in batch:
            yield dict(zip(columns, tup))


def _fetch_enrollment_student_summary(
    conn: duckdb.DuckDBPyConnection,
    course_ids: Sequence[int],
    display_programme: str,
) -> list[dict[str, Any]]:
    if not course_ids:
        return []
    courses = qualified_relation(courses_schema(), "dim_courses")
    enrollments = qualified_relation(staging_schema(), "stg_moodle_enrollments")
    clause, params = _course_id_filter(course_ids)
    return list(
        _iter_query_dicts(
            conn,
            f"""
            SELECT
                TRIM(e.user_idnumber) AS student_no,
                TRIM(e.user_fullname) AS student,
                TRIM(e.user_email) AS email,
                ? AS programme,
                COUNT(DISTINCT e.course_id) AS modules
            FROM {enrollments} AS e
            INNER JOIN {courses} AS dc ON dc.course_id = e.course_id
            WHERE {clause}
              AND e.primary_role = 'student'
              AND COALESCE(e.is_suspended, false) = false
              AND TRIM(COALESCE(e.user_idnumber, '')) <> ''
            GROUP BY 1, 2, 3, 4
            ORDER BY student
            """,
            [display_programme, *params],
        )
    )


def _fetch_module_summary(
    conn: duckdb.DuckDBPyConnection,
    course_ids: Sequence[int],
    display_programme: str,
) -> list[dict[str, Any]]:
    if not course_ids:
        return []
    courses = qualified_relation(courses_schema(), "dim_courses")
    enrollments = qualified_relation(staging_schema(), "stg_moodle_enrollments")
    clause, params = _course_id_filter(course_ids)
    return list(
        _iter_query_dicts(
            conn,
            f"""
            SELECT
                ? AS programme,
                TRIM(dc.course_shortname) AS module_code,
                TRIM(dc.course_fullname) AS module,
                COUNT(DISTINCT e.user_id) AS students
            FROM {courses} AS dc
            LEFT JOIN {enrollments} AS e
                ON e.course_id = dc.course_id
               AND e.primary_role = 'student'
               AND COALESCE(e.is_suspended, false) = false
            WHERE {clause}
            GROUP BY dc.course_id, dc.course_shortname, dc.course_fullname
            ORDER BY module_code
            """,
            [display_programme, *params],
        )
    )


def _iter_grade_rows(
    conn: duckdb.DuckDBPyConnection,
    course_ids: Sequence[int],
    display_programme: str,
) -> Iterator[dict[str, Any]]:
    if not course_ids:
        return
    courses = qualified_relation(courses_schema(), "dim_courses")
    grades = qualified_relation(staging_schema(), "stg_moodle_grades")
    items = qualified_relation(staging_schema(), "stg_moodle_grade_items")
    clause, params = _course_id_filter(course_ids)
    yield from _iter_query_dicts(
        conn,
        f"""
        SELECT
            TRIM(g.user_idnumber) AS student_no,
            TRIM(g.user_fullname) AS student,
            ? AS programme,
            TRIM(dc.course_shortname) AS module_code,
            TRIM(dc.course_fullname) AS module,
            TRIM(gi.grade_item_name) AS assessment,
            g.grade_raw,
            g.grade_percentage,
            g.submitted_at,
            g.graded_at
        FROM {grades} AS g
        INNER JOIN {courses} AS dc ON dc.course_id = g.course_id
        LEFT JOIN {items} AS gi
            ON gi.course_id = g.course_id AND gi.grade_item_id = g.grade_item_id
        WHERE {clause}
          AND TRIM(COALESCE(g.user_idnumber, '')) <> ''
        ORDER BY student, module_code, assessment
        """,
        [display_programme, *params],
    )


def build_workbook_offering_fallback(
    conn: duckdb.DuckDBPyConnection,
    category_name: str,
    course_prefix: str,
    output_dir: Path,
    display_programme_code: str,
    *,
    schema: str | None = None,
    programme_codes: Sequence[str] | None = None,
) -> Path:
    """Excel export for offerings with no gradebook mart rows."""
    from populate_gradebook_from_warehouse import (
        COURSE_NOTES_SHEET_TITLE,
        MART_SHEET_HEADERS,
        PROGRAMME_SUMMARY_HEADERS,
        add_header_only_sheet,
        append_data_row,
        finish_sheet,
        format_cell,
        write_gradebook_course_notes,
        write_headers,
        write_missed_assessments,
        write_submission_trends,
        write_upcoming_deadlines,
    )

    course_ids = course_ids_for_offering(conn, category_name, course_prefix)
    display = display_programme_code.strip().upper() or course_prefix.strip().upper()

    students = _fetch_enrollment_student_summary(conn, course_ids, display)
    modules = _fetch_module_summary(conn, course_ids, display)

    wb = Workbook(write_only=True)

    ws_prog = wb.create_sheet(title="Programme Summary"[:31])
    write_headers(ws_prog, PROGRAMME_SUMMARY_HEADERS)
    prog_values = [
        format_cell(display),
        format_cell(len(students)),
        "",
        format_cell(len(modules)),
        "",
        "",
        "",
        "",
        "",
    ]
    append_data_row(ws_prog, prog_values)
    finish_sheet(ws_prog, PROGRAMME_SUMMARY_HEADERS, 1)

    ws_stu = wb.create_sheet(title="Student Summary"[:31])
    stu_headers = ["Programme", "Student No", "Student", "Email", "Modules"]
    write_headers(ws_stu, stu_headers)
    for row in students:
        append_data_row(
            ws_stu,
            [
                format_cell(row.get("programme")),
                format_cell(row.get("student_no")),
                format_cell(row.get("student")),
                format_cell(row.get("email")),
                format_cell(row.get("modules")),
            ],
        )
    finish_sheet(ws_stu, stu_headers, len(students))

    ws_mod = wb.create_sheet(title="Module Summary"[:31])
    mod_headers = ["Programme", "Module Code", "Module", "Students"]
    write_headers(ws_mod, mod_headers)
    for row in modules:
        append_data_row(
            ws_mod,
            [
                format_cell(row.get("programme")),
                format_cell(row.get("module_code")),
                format_cell(row.get("module")),
                format_cell(row.get("students")),
            ],
        )
    finish_sheet(ws_mod, mod_headers, len(modules))

    mart_codes = list(programme_codes or [])
    mart_schema = schema

    def _mart_sheet(title: str, writer, header_key: str) -> None:
        if mart_schema and mart_codes:
            ws = wb.create_sheet(title=title[:31])
            writer(ws, conn, mart_schema, mart_codes, category_name)
        else:
            add_header_only_sheet(wb, title, MART_SHEET_HEADERS[header_key])

    _mart_sheet("Submission Trends", write_submission_trends, "Submission Trends")

    ws_det = wb.create_sheet(title="Student Assessment Detail"[:31])
    det_headers = [
        "Programme",
        "Student No",
        "Student",
        "Module Code",
        "Module",
        "Assessment",
        "Grade",
        "Grade %",
        "Submitted At",
        "Graded At",
    ]
    write_headers(ws_det, det_headers)
    grade_count = 0
    for row in _iter_grade_rows(conn, course_ids, display):
        append_data_row(
            ws_det,
            [
                format_cell(row.get("programme")),
                format_cell(row.get("student_no")),
                format_cell(row.get("student")),
                format_cell(row.get("module_code")),
                format_cell(row.get("module")),
                format_cell(row.get("assessment")),
                format_cell(row.get("grade_raw")),
                format_cell(row.get("grade_percentage")),
                format_cell(row.get("submitted_at")),
                format_cell(row.get("graded_at")),
            ],
        )
        grade_count += 1
    finish_sheet(ws_det, det_headers, grade_count)

    _mart_sheet("Missed Assessments", write_missed_assessments, "Missed Assessments")
    _mart_sheet("Upcoming Deadlines", write_upcoming_deadlines, "Upcoming Deadlines")
    ws_notes = wb.create_sheet(title=COURSE_NOTES_SHEET_TITLE[:31])
    write_gradebook_course_notes(
        ws_notes,
        conn,
        schema or "moodle_processed",
        mart_codes,
        category_name,
    )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_code = display.replace(" ", "_")
    out_path = output_dir / f"gradebook_{safe_code}_{timestamp}.xlsx"
    wb.save(out_path)
    return out_path.resolve()
