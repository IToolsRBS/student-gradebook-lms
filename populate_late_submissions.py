"""
Build Late Submissions export workbook from warehouse gradebook marts.

Filters (CLI; empty / omitted = select all for that dimension):
  category, programme, module, assessment type, assessment.
  Status defaults to submitted_late (effective status from assessment detail).

Sheets:
  1. Late Submissions Summary — one row per assessment (Total Late)
  2. Late Submission Details — same shape as Student Assessment Detail / activity details

Prints the absolute output path as the last stdout line (for future frontend wiring).
"""

from __future__ import annotations

import argparse
import gc
from datetime import datetime
from pathlib import Path
from typing import Any, Sequence

import duckdb
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from motherduck_client import connect_motherduck, gradebook_schema
from populate_activity_completion import (
    ALLOWED_STATUSES,
    DETAIL_HEADERS as ACTIVITY_DETAIL_HEADERS,
    _build_activity_filter_sql,
    _module_stats_lookup,
    _normalize_filter_values,
    _normalize_statuses,
    _suspended_by_module_from_detail,
    effective_completion_status,
    iter_filtered_assessment_rows,
)
from populate_gradebook_from_warehouse import (
    COLUMN_WIDTH_SAMPLE_ROWS,
    DEFAULT_SCHEMA,
    FETCH_CHUNK_SIZE,
    MAX_COLUMN_WIDTH,
    NOTE_FIELD_MAP,
    TABLE_ASSESSMENT,
    _mart_columns,
    _pick_first_mart_column,
    _update_col_widths,
    append_data_row,
    finish_sheet,
    format_cell,
    normalize_row,
    pick,
    write_headers,
)

DEFAULT_LATE_STATUSES: tuple[str, ...] = ("submitted_late",)

SUMMARY_HEADERS: list[str] = [
    "Category",
    "Programme",
    "Module Code",
    "Module",
    "Assessment",
    "Assessment Type",
    "Total Students In Module",
    "Total Students Suspended In Module",
    "Total Late",
]

DETAIL_HEADERS: list[str] = list(ACTIVITY_DETAIL_HEADERS)


def write_late_submissions_summary(
    ws: Worksheet,
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    *,
    category_names: Sequence[str],
    programme_codes: Sequence[str],
    modules: Sequence[str],
    assessment_types: Sequence[str],
    assessments: Sequence[str],
    statuses: Sequence[str],
    due_from: str | None = None,
    due_to: str | None = None,
) -> int:
    write_headers(ws, SUMMARY_HEADERS)
    widths = [max(12, min(len(h) + 2, MAX_COLUMN_WIDTH)) for h in SUMMARY_HEADERS]
    sample_remaining = [COLUMN_WIDTH_SAMPLE_ROWS]

    try:
        mart_cols = _mart_columns(conn, schema, TABLE_ASSESSMENT)
    except duckdb.CatalogException:
        finish_sheet(ws, len(SUMMARY_HEADERS), 0, widths)
        return 0

    category_col = _pick_first_mart_column(mart_cols, ("category_name",))
    programme_col = _pick_first_mart_column(
        mart_cols, ("programme", "course_prefix", "program_code")
    )
    module_code_col = _pick_first_mart_column(
        mart_cols, ("course_shortname", "module_code")
    )
    module_name_col = _pick_first_mart_column(mart_cols, ("course_fullname", "module"))
    assessment_col = _pick_first_mart_column(
        mart_cols, ("assessment_name", "assessment")
    )
    assessment_type_col = _pick_first_mart_column(mart_cols, ("assessment_type",))
    if not programme_col or not module_code_col or not assessment_col:
        finish_sheet(ws, len(SUMMARY_HEADERS), 0, widths)
        return 0

    select_parts = [
        (
            f'MAX("{category_col}") AS category_name'
            if category_col
            else "'' AS category_name"
        ),
        f'"{programme_col}" AS programme',
        f'"{module_code_col}" AS module_code',
        (
            f'MAX("{module_name_col}") AS module_name'
            if module_name_col
            else "'' AS module_name"
        ),
        f'"{assessment_col}" AS assessment_name',
        (
            f'MAX("{assessment_type_col}") AS assessment_type'
            if assessment_type_col
            else "'' AS assessment_type"
        ),
        "COUNT(*) AS total_late",
    ]
    group_parts = [f'"{programme_col}"', f'"{module_code_col}"', f'"{assessment_col}"']

    built = _build_activity_filter_sql(
        conn,
        schema,
        category_names=category_names,
        programme_codes=programme_codes,
        modules=modules,
        assessment_types=assessment_types,
        assessments=assessments,
        statuses=statuses,
        due_from=due_from,
        due_to=due_to,
        select_sql=", ".join(select_parts),
        group_by_sql=", ".join(group_parts),
        order_columns=["programme", "course_shortname", "assessment_name"],
    )
    if not built:
        finish_sheet(ws, len(SUMMARY_HEADERS), 0, widths)
        return 0

    query, params = built
    module_stats = _module_stats_lookup(
        conn,
        schema,
        category_names=category_names,
        programme_codes=programme_codes,
        modules=modules,
    )
    suspended_fallback = _suspended_by_module_from_detail(
        conn,
        schema,
        category_names=category_names,
        programme_codes=programme_codes,
        modules=modules,
    )

    result = conn.execute(query, params)
    columns = [str(desc[0]).lower() for desc in result.description]
    count = 0
    while True:
        batch = result.fetchmany(FETCH_CHUNK_SIZE)
        if not batch:
            break
        for tup in batch:
            row = normalize_row(dict(zip(columns, tup)))
            programme = str(pick(row, "programme") or "").strip().upper()
            module_code = str(pick(row, "module_code") or "").strip().upper()
            stats = module_stats.get((programme, module_code), {})
            students = pick(stats, "students", "total_students")
            suspended = pick(stats, "suspended_students", "suspended")
            if suspended == "" or suspended is None:
                suspended = suspended_fallback.get((programme, module_code), 0)
            module_name = pick(row, "module_name") or pick(
                stats, "module_name", "module"
            )
            category = pick(row, "category_name") or pick(stats, "category_name")
            values = [
                format_cell(category),
                format_cell(pick(row, "programme")),
                format_cell(pick(row, "module_code")),
                format_cell(module_name),
                format_cell(pick(row, "assessment_name")),
                format_cell(pick(row, "assessment_type")),
                format_cell(students if students != "" else 0),
                format_cell(suspended if suspended != "" else 0),
                format_cell(pick(row, "total_late")),
            ]
            _update_col_widths(widths, values, sample_remaining)
            append_data_row(ws, values)
            count += 1

    finish_sheet(ws, len(SUMMARY_HEADERS), count, widths)
    return count


def write_late_submission_details(
    ws: Worksheet,
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    *,
    category_names: Sequence[str],
    programme_codes: Sequence[str],
    modules: Sequence[str],
    assessment_types: Sequence[str],
    assessments: Sequence[str],
    statuses: Sequence[str],
    due_from: str | None = None,
    due_to: str | None = None,
) -> int:
    write_headers(ws, DETAIL_HEADERS)
    widths = [max(12, min(len(h) + 2, MAX_COLUMN_WIDTH)) for h in DETAIL_HEADERS]
    sample_remaining = [COLUMN_WIDTH_SAMPLE_ROWS]
    count = 0

    for row in iter_filtered_assessment_rows(
        conn,
        schema,
        category_names=category_names,
        programme_codes=programme_codes,
        modules=modules,
        assessment_types=assessment_types,
        assessments=assessments,
        statuses=statuses,
        due_from=due_from,
        due_to=due_to,
        order_columns=[
            "category_name",
            "programme",
            "course_prefix",
            "course_shortname",
            "assessment_name",
            "student_no",
        ],
    ):
        submitted = pick(
            row,
            "grade_submitted_at",
            "last_attempt_at",
            "submitted_at",
            "graded_at",
        )
        status_val = effective_completion_status(row)
        if not submitted and status_val not in ALLOWED_STATUSES:
            submitted = ""

        values = [
            format_cell(pick(row, "category_name")),
            format_cell(pick(row, "programme", "course_prefix", "program_code")),
            format_cell(pick(row, "student_no")),
            format_cell(pick(row, "user_fullname")),
            format_cell(pick(row, "user_email")),
            format_cell(pick(row, "course_shortname", "module_code")),
            format_cell(pick(row, "course_fullname", "module")),
            format_cell(pick(row, "assessment_name", "assessment")),
            format_cell(pick(row, "assessment_type")),
            format_cell(pick(row, "due_at", "effective_deadline_at")),
            format_cell(submitted),
            format_cell(status_val),
            format_cell(pick(row, "mark_status")),
            format_cell(pick(row, "grade_raw")),
            format_cell(pick(row, "max_grade")),
            *[format_cell(pick(row, *aliases)) for _, aliases in NOTE_FIELD_MAP],
        ]
        _update_col_widths(widths, values, sample_remaining)
        append_data_row(ws, values)
        count += 1

    finish_sheet(ws, len(DETAIL_HEADERS), count, widths)
    return count


def build_workbook(
    conn: duckdb.DuckDBPyConnection,
    schema: str,
    output_dir: Path,
    *,
    category_names: Sequence[str],
    programme_codes: Sequence[str],
    modules: Sequence[str],
    assessment_types: Sequence[str],
    assessments: Sequence[str],
    statuses: Sequence[str],
    due_from: str | None = None,
    due_to: str | None = None,
) -> Path:
    wb = Workbook(write_only=True)

    ws_summary = wb.create_sheet(title="Late Submissions Summary"[:31])
    write_late_submissions_summary(
        ws_summary,
        conn,
        schema,
        category_names=category_names,
        programme_codes=programme_codes,
        modules=modules,
        assessment_types=assessment_types,
        assessments=assessments,
        statuses=statuses,
        due_from=due_from,
        due_to=due_to,
    )
    gc.collect()

    ws_detail = wb.create_sheet(title="Late Submission Details"[:31])
    write_late_submission_details(
        ws_detail,
        conn,
        schema,
        category_names=category_names,
        programme_codes=programme_codes,
        modules=modules,
        assessment_types=assessment_types,
        assessments=assessments,
        statuses=statuses,
        due_from=due_from,
        due_to=due_to,
    )
    gc.collect()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if len(programme_codes) == 1:
        file_code = programme_codes[0]
    elif programme_codes:
        joined = "_".join(programme_codes)
        file_code = joined if len(joined) <= 48 else f"batch_{len(programme_codes)}prog"
    else:
        file_code = "all_programmes"
    safe_code = file_code.replace(" ", "_")
    out_path = output_dir / f"late_submissions_{safe_code}_{timestamp}.xlsx"
    wb.save(out_path)
    return out_path.resolve()


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Export Late Submissions Excel from warehouse marts. "
            "Omit a filter (or pass *) for select-all on that dimension. "
            "Status defaults to submitted_late."
        )
    )
    parser.add_argument(
        "--category-name",
        action="append",
        dest="category_names",
        default=None,
        help="Category / intake name (repeatable; omit = all)",
    )
    parser.add_argument(
        "--programme-code",
        action="append",
        dest="programme_codes",
        default=None,
        help="Programme code (repeatable; omit = all)",
    )
    parser.add_argument(
        "--module",
        action="append",
        dest="modules",
        default=None,
        help="Module code / course shortname (repeatable; omit = all)",
    )
    parser.add_argument(
        "--assessment-type",
        action="append",
        dest="assessment_types",
        default=None,
        help="Assessment type (repeatable; omit = all)",
    )
    parser.add_argument(
        "--assessment",
        action="append",
        dest="assessments",
        default=None,
        help="Assessment name (repeatable; omit = all)",
    )
    parser.add_argument(
        "--status",
        action="append",
        dest="statuses",
        default=None,
        help=(
            "Completion status filter (repeatable). "
            f"Default: {', '.join(DEFAULT_LATE_STATUSES)}. "
            f"Allowed: {', '.join(ALLOWED_STATUSES)}."
        ),
    )
    parser.add_argument(
        "--due-from",
        default=None,
        help="Optional due date lower bound (YYYY-MM-DD)",
    )
    parser.add_argument(
        "--due-to",
        default=None,
        help="Optional due date upper bound (YYYY-MM-DD)",
    )
    parser.add_argument(
        "--warehouse-schema",
        default=None,
        help="Schema for gradebook marts (default: moodle_processed)",
    )
    parser.add_argument(
        "--output-dir",
        default=".",
        help="Directory for the generated workbook",
    )
    args = parser.parse_args()

    category_names = _normalize_filter_values(args.category_names)
    programme_codes = [
        c.strip().upper() for c in _normalize_filter_values(args.programme_codes)
    ]
    modules = [m.strip().upper() for m in _normalize_filter_values(args.modules)]
    assessment_types = [
        t.strip().upper() for t in _normalize_filter_values(args.assessment_types)
    ]
    assessments = [
        a.strip().upper() for a in _normalize_filter_values(args.assessments)
    ]
    # Default to late-only; if --status is passed, honour the activity allowlist.
    if args.statuses:
        statuses = _normalize_statuses(args.statuses)
    else:
        statuses = list(DEFAULT_LATE_STATUSES)
    due_from = (args.due_from or "").strip() or None
    due_to = (args.due_to or "").strip() or None

    schema = args.warehouse_schema or gradebook_schema() or DEFAULT_SCHEMA
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    conn = connect_motherduck()
    try:
        out_path = build_workbook(
            conn,
            schema,
            output_dir,
            category_names=category_names,
            programme_codes=programme_codes,
            modules=modules,
            assessment_types=assessment_types,
            assessments=assessments,
            statuses=statuses,
            due_from=due_from,
            due_to=due_to,
        )
    except Exception:
        import traceback

        traceback.print_exc()
        raise
    finally:
        conn.close()

    print(out_path)


if __name__ == "__main__":
    main()
